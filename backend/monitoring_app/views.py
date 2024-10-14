import os
import json
import time
import logging
import zipfile
import datetime
from tempfile import NamedTemporaryFile
from concurrent.futures import ThreadPoolExecutor

from drf_yasg import openapi
from django.conf import settings
from django.utils import timezone
from rest_framework import status
from openpyxl import load_workbook
from django.contrib import messages
from django.core.cache import caches
from django.http import HttpResponse
from django.db.models import Count, Q
from django.views.generic import View
from django.contrib.auth.models import User
from django.core.files.base import ContentFile
from drf_yasg.utils import swagger_auto_schema
from django.db import IntegrityError, transaction
from rest_framework.pagination import PageNumberPagination
from django.contrib.auth import authenticate, login, logout
from django.shortcuts import get_object_or_404, redirect, render
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import (
    AllowAny,
    IsAdminUser,
    IsAuthenticated,
)
from rest_framework.views import APIView
from rest_framework.response import Response

from celery.result import AsyncResult
from monitoring_app import models, permissions, serializers, utils, tasks

logger = logging.getLogger(__name__)


class StaffAttendancePagination(PageNumberPagination):
    page_size = 5000
    page_size_query_param = "page_size"
    max_page_size = 20000


Cache = caches["default"]
token_param_config = openapi.Parameter(
    "Authorization",
    in_=openapi.IN_HEADER,
    description="Token [access_token]",
    type=openapi.TYPE_STRING,
)


def get_cache(
    key: str,
    query: callable = lambda: any,
    timeout: int = 10,
    cache: any = Cache,
) -> any:
    """
    Получает данные из кэша по указанному ключу `key`.

    Args:
        key (str): Строковый ключ для доступа к данным в кэше.
        query (callable, optional): Функция, вызываемая для получения данных в случае их отсутствия в кэше.
            По умолчанию используется `lambda: any`, возвращающая всегда `True`.
        timeout (int, optional): Время жизни данных в кэше в секундах. По умолчанию: 10 секунд.
        cache (any, optional): Объект кэша, используемый для хранения данных. По умолчанию: `Cache`.

    Returns:
        any: Возвращает данные из кэша, если они есть, иначе данные, полученные из запроса.

    Examples:
        >>> get_cache("my_data_key")
    """
    data = cache.get(key)
    if data is None:
        data = query()
        cache.set(key, data, timeout)
    return data


@permission_classes([AllowAny])
def home(request):
    return render(
        request,
        "index.html",
    )


@permission_classes([AllowAny])
def react_app(request):
    def render_react_app():
        try:
            return render(request, "index.html")
        except Exception as error:
            logger.error(f"React App {str(error)}")
            return None

    with ThreadPoolExecutor(max_workers=3) as executor:
        future = executor.submit(render_react_app)
        response = future.result()

    if response is None:
        return HttpResponse("Error loading React app", status=500)

    return response


class StaffAttendanceStatsView(APIView):
    """
    Представление для получения статистики о посещаемости персонала.

    Это представление фильтрует данные, чтобы включать только сотрудников, относящихся к отделу с ID 4958.

    Параметры запроса:
        date (str): Дата, для которой запрашивается статистика посещаемости, в формате 'YYYY-MM-DD'. По умолчанию используется текущая дата.
        pin (str): ПИН-код сотрудника для получения конкретной статистики посещаемости.

    Возвращает:
        JSON-ответ, содержащий:
            department_name (str): Название отдела.
            total_staff_count (int): Общее количество сотрудников.
            present_staff_count (int): Количество присутствующих сотрудников.
            absent_staff_count (int): Количество отсутствующих сотрудников.
            present_between_9_to_18 (int): Количество сотрудников, присутствующих с 08:00 до 19:00.
            present_data (list): Список словарей с информацией о присутствующих сотрудниках, включая ПИН, имя, количество минут присутствия и индивидуальный процент.
            absent_data (list): Список словарей с информацией об отсутствующих сотрудниках, включая ПИН и имя.
            data_for_date (str): Дата, за которую предоставлены данные, в формате 'YYYY-MM-DD'.

    Примеры:
        GET /api/attendance/stats/?date=2024-07-20
        GET /api/attendance/stats/?pin=123456

    Примечание:
        Ответ кэшируется на 1 час, а информация о государственных праздниках кэшируется на 1 минуту.
    """

    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        operation_summary="Получить список людей об их присутствии",
        operation_description="View для получения статистики о посещаемости персонала.",
        responses={
            200: openapi.Response(
                description="Successful response",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        "department_name": openapi.Schema(type=openapi.TYPE_STRING),
                        "total_staff_count": openapi.Schema(type=openapi.TYPE_INTEGER),
                        "present_staff_count": openapi.Schema(
                            type=openapi.TYPE_INTEGER
                        ),
                        "absent_staff_count": openapi.Schema(type=openapi.TYPE_INTEGER),
                        "present_between_9_to_18": openapi.Schema(
                            type=openapi.TYPE_INTEGER
                        ),
                        "present_data": openapi.Schema(
                            type=openapi.TYPE_ARRAY,
                            items=openapi.Schema(
                                type=openapi.TYPE_OBJECT,
                                properties={
                                    "staff_pin": openapi.Schema(
                                        type=openapi.TYPE_STRING
                                    ),
                                    "name": openapi.Schema(type=openapi.TYPE_STRING),
                                    "minutes_present": openapi.Schema(
                                        type=openapi.TYPE_NUMBER
                                    ),
                                    "individual_percentage": openapi.Schema(
                                        type=openapi.TYPE_NUMBER
                                    ),
                                },
                            ),
                        ),
                        "absent_data": openapi.Schema(
                            type=openapi.TYPE_ARRAY,
                            items=openapi.Schema(
                                type=openapi.TYPE_OBJECT,
                                properties={
                                    "staff_pin": openapi.TYPE_STRING,
                                    "name": openapi.Schema(type=openapi.TYPE_STRING),
                                },
                            ),
                        ),
                        "data_for_date": openapi.Schema(type=openapi.TYPE_STRING),
                    },
                ),
            ),
            404: "Not Found",
            500: "Internal Server Error",
        },
        manual_parameters=[
            openapi.Parameter(
                "date",
                openapi.IN_QUERY,
                description="Date in 'YYYY-MM-DD' format.",
                type=openapi.TYPE_STRING,
            ),
            openapi.Parameter(
                "pin",
                openapi.IN_QUERY,
                description="Staff PIN.",
                type=openapi.TYPE_STRING,
            ),
        ],
    )
    def get(self, request):
        logger.info("Received request for staff attendance stats.")

        date_param = request.query_params.get(
            "date", timezone.now().date().strftime("%Y-%m-%d")
        )
        date_param = datetime.datetime.strptime(date_param, "%Y-%m-%d").date()
        pin_param = request.query_params.get("pin", None)

        logger.debug(f"Parsed date_param: {date_param}, pin_param: {pin_param}")

        try:
            target_date = self.get_last_working_day(date_param)
            next_date = target_date + datetime.timedelta(days=1)
            cache_key = f"staff_attendance_stats_{target_date}_{pin_param}"

            logger.debug(f"Generated cache_key: {cache_key}")

            cached_data = get_cache(
                cache_key,
                query=lambda: self.query_data(target_date, next_date, pin_param),
                timeout=1 * 5 * 60,
            )

            logger.info("Successfully retrieved staff attendance data.")
            return Response(cached_data)

        except Exception as e:
            logger.error(f"Error while processing request: {str(e)}")
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def get_last_working_day(self, date):
        """
        Определение последнего рабочего дня, с учетом выходных и государственных праздников.

        Args:
            date (datetime.date): Дата, для которой нужно найти последний рабочий день.

        Returns:
            datetime.date: Последний рабочий день.
        """
        logger.debug(f"Calculating last working day for date: {date}")

        holidays = get_cache(
            "public_holidays",
            query=lambda: list(models.PublicHoliday.objects.all()),
            timeout=10 * 6,
        )
        holiday_dates = {holiday.date: holiday.is_working_day for holiday in holidays}

        if date.weekday() == 5:
            logger.debug(f"Today is Saturday ({date}), checking Friday.")
            date -= datetime.timedelta(days=1)
        elif date.weekday() == 6:
            logger.debug(f"Today is Sunday ({date}), checking Friday.")
            date -= datetime.timedelta(days=2)
        elif date.weekday() == 0:
            logger.debug(f"Today is Monday ({date}), checking Friday.")
            date -= datetime.timedelta(days=3)

        if date in holiday_dates and not holiday_dates[date]:
            logger.debug(
                f"Friday {date} is a holiday or non-working day, searching for previous working day."
            )
            while date in holiday_dates and not holiday_dates[date]:
                logger.debug(
                    f"{date} is a holiday or non-working day, moving to previous day."
                )
                date -= datetime.timedelta(days=1)

        while date.weekday() >= 5 or (
            date in holiday_dates and not holiday_dates[date]
        ):
            logger.debug(
                f"{date} is a weekend or non-working day, moving to previous day."
            )
            date -= datetime.timedelta(days=1)

        logger.debug(f"Last working day determined: {date}")
        return date

    def query_data(
        self,
        target_date: datetime.date,
        next_date: datetime.date,
        pin_param: str | None,
    ) -> dict:
        """
        Запрашивает данные по целевой дате и отделу (или сотруднику).

        Args:
            target_date (datetime.date): Целевая дата.
            next_date (datetime.date): Следующая дата после целевой.
            pin_param (str, optional): ID родительского или дочернего отдела.

        Returns:
            dict: Данные о сотрудниках, их посещаемости и статистике.
        """
        logger.info(
            f"Querying data for target_date: {target_date}, pin_param: {pin_param}"
        )

        department_name = "Unknown Department"
        staff_queryset = None

        parent_department = models.ParentDepartment.objects.filter(id=pin_param).first()
        child_department = models.ChildDepartment.objects.filter(id=pin_param).first()

        match (parent_department, child_department):
            case (parent, None) if parent:
                staff_queryset = models.Staff.objects.filter(
                    department__parent_id=parent.id
                ).select_related("department")
                department_name = parent.name
            case (None, child) if child:
                staff_queryset = models.Staff.objects.filter(
                    department=child
                ).select_related("department")
                department_name = child.name
            case _:
                staff_queryset = models.Staff.objects.filter(
                    department__parent_id="4958"
                ).select_related("department")
                department_name = (
                    staff_queryset.first().department.parent.name
                    if staff_queryset.exists()
                    else "Unknown Department"
                )

        target_date_for_filter = target_date + datetime.timedelta(days=1)
        staff_attendance_queryset = models.StaffAttendance.objects.filter(
            date_at=target_date_for_filter, staff__in=staff_queryset
        ).select_related("staff")

        total_staff_count = staff_queryset.count()
        present_staff = staff_attendance_queryset.exclude(first_in__isnull=True)
        present_staff_pins = set(present_staff.values_list("staff__pin", flat=True))
        absent_staff_count = total_staff_count - len(present_staff_pins)
        present_between_9_to_18 = present_staff.filter(
            first_in__time__range=["08:00", "19:00"]
        ).count()

        present_data, absent_data = self.get_attendance_data(
            staff_queryset, present_staff_pins, present_staff
        )

        logger.info(f"Data query successful for department: {department_name}")

        return {
            "department_name": department_name,
            "total_staff_count": total_staff_count,
            "present_staff_count": len(present_data),
            "absent_staff_count": absent_staff_count,
            "present_between_9_to_18": present_between_9_to_18,
            "present_data": present_data,
            "absent_data": absent_data,
            "data_for_date": target_date.strftime("%Y-%m-%d"),
        }

    def get_attendance_data(self, staff_queryset, present_staff_pins, present_staff):
        logger.debug("Generating attendance data.")
        present_data = []
        absent_data = []
        total_minutes = 8 * 60
        for staff in staff_queryset:
            if staff.pin in present_staff_pins:
                attendance = present_staff.get(staff__pin=staff.pin)
                minutes_present = (
                    (attendance.last_out - attendance.first_in).total_seconds() / 60
                    if attendance.last_out
                    else 0
                )
                individual_percentage = (minutes_present / total_minutes) * 100
                present_data.append(
                    {
                        "staff_pin": staff.pin,
                        "name": f"{staff.surname} {staff.name}",
                        "minutes_present": round(minutes_present, 2),
                        "individual_percentage": round(individual_percentage, 2),
                    }
                )
            else:
                absent_data.append(
                    {
                        "staff_pin": staff.pin,
                        "name": f"{staff.surname} {staff.name}",
                    }
                )

        logger.info("Attendance data generation complete.")
        return present_data, absent_data


@swagger_auto_schema(
    method="GET",
    operation_summary="Получить данные о локациях и сотрудниках по турникетам",
    operation_description="Метод возвращает данные по количеству сотрудников, зарегистрированных в разных локациях на основе зоны турникетов.",
    responses={
        200: openapi.Response(
            description="Успешный запрос. Возвращается список данных о локациях с количеством сотрудников.",
            schema=openapi.Schema(
                type=openapi.TYPE_ARRAY,
                items=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        "name": openapi.Schema(
                            type=openapi.TYPE_STRING,
                            description="Название здания или локации.",
                        ),
                        "lat": openapi.Schema(
                            type=openapi.TYPE_NUMBER,
                            format="float",
                            description="Широта координаты.",
                        ),
                        "lng": openapi.Schema(
                            type=openapi.TYPE_NUMBER,
                            format="float",
                            description="Долгота координаты.",
                        ),
                        "employees": openapi.Schema(
                            type=openapi.TYPE_INTEGER,
                            description="Количество сотрудников.",
                        ),
                    },
                ),
            ),
        ),
        404: openapi.Response(
            description="Данные не найдены для запрашиваемой даты.",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "error": openapi.Schema(type=openapi.TYPE_STRING),
                },
            ),
        ),
        500: openapi.Response(
            description="Внутренняя ошибка сервера.",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "error": openapi.Schema(type=openapi.TYPE_STRING),
                },
            ),
        ),
    },
)
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def map_location(request):
    """
    Возвращает данные о локациях и количестве сотрудников в них на основе турникетов за указанную дату.

    Args:
        request (HttpRequest): HTTP-запрос с необязательным параметром `date_at`, который используется для фильтрации данных.

    Returns:
        JsonResponse: JSON-ответ с данными о локациях, если запрос успешен, либо сообщение об ошибке.

    Raises:
        KeyError: Если в данных присутствует неверный ключ.
        ValueError: Если параметры запроса содержат неверные значения.
        Exception: Для обработки любых других непредвиденных исключений.
    """
    try:
        zone_mapping = {
            "Главный Корпус (Абылайхана)": {
                "lat": 43.2644734,
                "lng": 76.9393907,
                "areas": [
                    "Абылайхана турникет",
                    "вход в 8 этаж",
                    "военные 3 этаж",
                    "лифтовые с 1 по 7",
                    "выход ЦОС",
                ],
            },
            "Второй Корпус (Торекулова)": {
                "lat": 43.2655509,
                "lng": 76.9299558,
                "areas": ["Торекулва турникет"],
            },
            "Третий Корпус (Карасай батыра)": {
                "lat": 43.251326,
                "lng": 76.9349449,
                "areas": ["карасай батыра турникет"],
            },
        }

        date_at = request.GET.get("date_at", None)
        if not date_at:
            logger.warning("No date_at parameter provided, using current date.")
            date_at = timezone.now().date()
        else:
            logger.info(f"Date parameter provided: {date_at}")

        cache_key = f"map_location_{date_at}"

        cached_data = get_cache(
            cache_key,
            query=lambda: generate_map_data(zone_mapping, date_at),
            timeout=1 * 60 * 60,
            cache=Cache,
        )

        return Response(cached_data, status=status.HTTP_200_OK)

    except KeyError as ke:
        logger.error(f"KeyError encountered: {str(ke)}", exc_info=True)
        return Response(
            {"error": "Internal error: data inconsistency."},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    except ValueError as ve:
        logger.error(f"ValueError encountered: {str(ve)}", exc_info=True)
        return Response(
            {"error": "Invalid input provided."}, status=status.HTTP_400_BAD_REQUEST
        )

    except Exception as e:
        logger.critical(f"Critical error in map_location: {str(e)}", exc_info=True)
        return Response(
            {"error": "A critical error occurred. Please try again later."},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


def generate_map_data(zone_mapping, date_at):
    """
    Генерирует данные по локациям и количеству сотрудников для указанной даты.

    Args:
        zone_mapping (dict): Словарь с информацией о зонах.
        date_at (str): Дата для фильтрации данных.

    Returns:
        list: Список данных по локациям и количеству сотрудников.
    """
    try:
        attendance_data = (
            models.StaffAttendance.objects.filter(
                date_at=date_at, first_in__isnull=False
            )
            .values("area_name_in")
            .annotate(employees=Count("staff"))
        )
        logger.info(f"Attendance data retrieved for date: {date_at}")
    except models.StaffAttendance.DoesNotExist:
        logger.error(f"No attendance records found for date: {date_at}")
        return []

    zone_result = {}

    for record in attendance_data:
        area_name = record["area_name_in"]
        matched = False

        for zone_name, zone_info in zone_mapping.items():
            if area_name in zone_info["areas"]:
                if zone_name in zone_result:
                    zone_result[zone_name]["employees"] += record["employees"]
                    logger.info(
                        f"Updated employee count for {zone_name}: {zone_result[zone_name]['employees']}"
                    )
                else:
                    zone_result[zone_name] = {
                        "name": zone_name,
                        "lat": zone_info["lat"],
                        "lng": zone_info["lng"],
                        "employees": record["employees"],
                    }
                    logger.info(
                        f"Added new zone entry: {zone_name} with {record['employees']} employees."
                    )
                matched = True
                break

        if not matched:
            logger.warning(
                f"Unknown area_name_in '{area_name}' found in attendance data. Skipped."
            )

    result_list = list(zone_result.values())

    main_building = next(
        (item for item in result_list if item["name"] == "Главный Корпус (Абылайхана)"),
        None,
    )

    if main_building:
        result_list.remove(main_building)
        result_list.insert(0, main_building)

    return result_list


@swagger_auto_schema(
    method="GET",
    operation_summary="Получить ID всех родительских департаментов",
    operation_description="Метод для получения списка всех родительских департаментов.",
    responses={
        200: openapi.Response(
            description="Успешный запрос. Возвращается список ID родительских департаментов.",
            schema=openapi.Schema(
                type=openapi.TYPE_ARRAY,
                items=openapi.Schema(
                    type=openapi.TYPE_INTEGER,
                    description="ID родительского департамента.",
                ),
            ),
        ),
        404: openapi.Response(
            description="Не удалось найти департаменты",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "error": openapi.Schema(type=openapi.TYPE_STRING),
                },
            ),
        ),
    },
)
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_parent_id(request):
    """
    Получить ID всех родительских департаментов.

    Этот метод возвращает список всех ID родительских департаментов.

    Возвращаемые данные:
    - Список ID родительских департаментов.

    Возможные ошибки:
    - 404: Департаменты не найдены.

    Пример ответа:
    - 200: [1, 2, 3, ...]
    - 404: {"error": "Не удалось найти департаменты."}
    """
    logger.info("Request received to get parent department IDs.")

    try:
        parent_departments = models.ParentDepartment.objects.all()
        parent_ids = [department.id for department in parent_departments]
        logger.info(f"Found {len(parent_ids)} parent departments.")
        return Response(data=parent_ids, status=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f"Error retrieving parent department IDs: {str(e)}")
        return Response(data={"error": str(e)}, status=status.HTTP_404_NOT_FOUND)


@swagger_auto_schema(
    method="GET",
    operation_summary="Сводная информация о департаменте",
    operation_description="Метод для получения сводной информации о департаменте и его дочерних подразделениях с количеством сотрудников.",
    responses={
        200: openapi.Response(
            description="Успешный запрос. Возвращается сводная информация о департаменте и его дочерних подразделениях.",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "name": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Название департамента.",
                    ),
                    "date_of_creation": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        format="date-time",
                        description="Дата создания департамента.",
                    ),
                    "child_departments": openapi.Schema(
                        type=openapi.TYPE_ARRAY,
                        items=openapi.Schema(
                            type=openapi.TYPE_OBJECT,
                            properties={
                                "child_id": openapi.Schema(
                                    type=openapi.TYPE_STRING,
                                    description="ID дочернего подразделения.",
                                ),
                                "name": openapi.Schema(
                                    type=openapi.TYPE_STRING,
                                    description="Название дочернего подразделения.",
                                ),
                                "date_of_creation": openapi.Schema(
                                    type=openapi.TYPE_STRING,
                                    format="date-time",
                                    description="Дата создания дочернего подразделения.",
                                ),
                                "parent": openapi.Schema(
                                    type=openapi.TYPE_STRING,
                                    description="ID родительского департамента.",
                                ),
                            },
                        ),
                        description="Список дочерних подразделений департамента.",
                    ),
                    "total_staff_count": openapi.Schema(
                        type=openapi.TYPE_INTEGER,
                        description="Общее количество сотрудников в департаменте и его дочерних подразделениях.",
                    ),
                },
            ),
        ),
        404: "Департамент не найден.",
    },
)
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def department_summary(request, parent_department_id):
    cache_key = f"department_summary_{parent_department_id}"
    logger.info(
        f"Request received for department summary with ID {parent_department_id}"
    )

    if not models.ChildDepartment.objects.filter(id=parent_department_id).exists():
        logger.warning(f"Department with ID {parent_department_id} not found")
        return Response(
            status=status.HTTP_404_NOT_FOUND,
            data={"message": f"Department with ID {parent_department_id} not found"},
        )

    try:

        def calculate_staff_count(department):
            logger.debug(f"Calculating staff count for department ID {department.id}")
            child_departments = models.ChildDepartment.objects.filter(parent=department)
            staff_count = (
                child_departments.aggregate(total_staff=Count("staff"))["total_staff"]
                or 0
            )

            for child_dept in child_departments:
                staff_count += calculate_staff_count(child_dept)

            logger.debug(
                f"Total staff count for department ID {department.id} is {staff_count}"
            )
            return staff_count

        parent_department = get_object_or_404(
            models.ChildDepartment, id=parent_department_id
        )
        logger.info(
            f"Department found: {parent_department.name} (ID: {parent_department_id})"
        )
        parent_department_id = str(parent_department_id).zfill(5)
        total_staff_count = calculate_staff_count(parent_department)

        child_departments_data = models.ChildDepartment.objects.filter(
            parent=parent_department
        )
        child_departments_data_serialized = serializers.ChildDepartmentSerializer(
            child_departments_data, many=True
        ).data

        data = {
            "name": parent_department.name,
            "date_of_creation": parent_department.date_of_creation,
            "child_departments": child_departments_data_serialized,
            "total_staff_count": total_staff_count,
        }

        logger.debug(f"Caching department summary data with key: {cache_key}")
        cached_data = get_cache(
            cache_key, query=lambda: data, timeout=1 * 5, cache=Cache
        )

        logger.info(f"Returning summary data for department ID {parent_department_id}")
        return Response(cached_data, status=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f"Error while generating department summary: {str(e)}")
        return Response(data={"message": str(e)}, status=status.HTTP_404_NOT_FOUND)


@swagger_auto_schema(
    method="get",
    operation_summary="Получить описание подотдела",
    operation_description="Получите подробную информацию о подотделе и его сотрудниках.",
    manual_parameters=[
        openapi.Parameter(
            name="child_department_id",
            in_=openapi.IN_PATH,
            type=openapi.TYPE_INTEGER,
            description="ID подотдела",
            required=True,
        ),
    ],
    responses={
        200: openapi.Response(
            description="Сведения о подотделе и данные о персонале",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "child_department": openapi.Schema(
                        type=openapi.TYPE_OBJECT,
                        properties={
                            "id": openapi.Schema(type=openapi.TYPE_INTEGER),
                            "name": openapi.Schema(type=openapi.TYPE_STRING),
                            "date_of_creation": openapi.Schema(
                                type=openapi.TYPE_STRING,
                                format=openapi.FORMAT_DATETIME,
                            ),
                            "parent": openapi.Schema(type=openapi.TYPE_INTEGER),
                        },
                    ),
                    "staff_count": openapi.Schema(type=openapi.TYPE_INTEGER),
                    "staff_data": openapi.Schema(
                        type=openapi.TYPE_OBJECT,
                        additional_properties=openapi.Schema(
                            type=openapi.TYPE_OBJECT,
                            properties={
                                "FIO": openapi.Schema(type=openapi.TYPE_STRING),
                                "date_of_creation": openapi.Schema(
                                    type=openapi.TYPE_STRING,
                                    format=openapi.FORMAT_DATETIME,
                                ),
                                "avatar": openapi.Schema(type=openapi.TYPE_STRING),
                                "positions": openapi.Schema(
                                    type=openapi.TYPE_ARRAY,
                                    items=openapi.Schema(type=openapi.TYPE_STRING),
                                ),
                            },
                        ),
                    ),
                },
            ),
        ),
        404: "Not Found: Если подотдела не существует.",
    },
)
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def child_department_detail(request, child_department_id):
    """
    Получите подробную информацию о дочернем отделе вместе с его сотрудниками.

    Args:
    запрос: объект запроса.
    child_department_id (int): идентификатор дочернего отдела, который требуется получить.

    Returns:
    Ответ: ответ, содержащий сведения о дочернем отделе и данные о сотрудниках.

    Raises:
    Http404: Если дочерний отдел не существует.
    """
    logger.info(
        f"Request received for child department detail with ID {child_department_id}"
    )

    try:
        child_department = models.ChildDepartment.objects.get(id=child_department_id)
        logger.info(
            f"Found child department: {child_department.name} (ID: {child_department_id})"
        )
    except models.ChildDepartment.DoesNotExist:
        logger.warning(f"Child department with ID {child_department_id} not found")
        return Response(status=status.HTTP_404_NOT_FOUND)

    all_departments = [child_department] + child_department.get_all_child_departments()
    staff_in_department = models.Staff.objects.filter(department__in=all_departments)
    logger.debug(
        f"Found {staff_in_department.count()} staff members in child department ID {child_department_id}"
    )

    staff_data = {}
    for staff_member in staff_in_department:
        if staff_member.surname == "Нет фамилии":
            fio = staff_member.name
        else:
            fio = f"{staff_member.surname} {staff_member.name}"

        staff_data[staff_member.pin] = {
            "FIO": fio,
            "date_of_creation": staff_member.date_of_creation,
            "avatar": (staff_member.avatar.url if staff_member.avatar else None),
            "positions": [position.name for position in staff_member.positions.all()],
        }
        logger.debug(f"Processed staff member: {fio} (PIN: {staff_member.pin})")

    sorted_staff_data = dict(
        sorted(staff_data.items(), key=lambda item: item[1]["FIO"])
    )
    logger.info(f"Sorted staff data for child department ID {child_department_id}")

    data = {
        "child_department": serializers.ChildDepartmentSerializer(
            child_department
        ).data,
        "staff_count": staff_in_department.count(),
        "staff_data": sorted_staff_data,
    }

    logger.info(
        f"Returning detailed data for child department ID {child_department_id}"
    )
    return Response(data, status=status.HTTP_200_OK)


@swagger_auto_schema(
    method="GET",
    operation_summary="Получить информацию о сотруднике",
    operation_description="Получение подробной информации о сотруднике, включая данные о посещаемости, заработной плате и типе контракта.",
    manual_parameters=[
        openapi.Parameter(
            name="staff_pin",
            in_=openapi.IN_PATH,
            type=openapi.TYPE_STRING,
            required=True,
            description="Уникальный идентификатор сотрудника (PIN)",
        ),
        openapi.Parameter(
            name="start_date",
            in_=openapi.IN_QUERY,
            type=openapi.TYPE_STRING,
            required=False,
            description="Дата начала периода для фильтрации данных о посещаемости (формат: YYYY-MM-DD)",
        ),
        openapi.Parameter(
            name="end_date",
            in_=openapi.IN_QUERY,
            type=openapi.TYPE_STRING,
            required=False,
            description="Дата окончания периода для фильтрации данных о посещаемости (формат: YYYY-MM-DD)",
        ),
    ],
    responses={
        200: openapi.Response(
            description="Данные сотрудника успешно получены",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "name": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Имя сотрудника",
                    ),
                    "surname": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Фамилия сотрудника",
                    ),
                    "positions": openapi.Schema(
                        type=openapi.TYPE_ARRAY,
                        items=openapi.Schema(type=openapi.TYPE_STRING),
                        description="Список должностей сотрудника",
                    ),
                    "avatar": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        format=openapi.FORMAT_URI,
                        nullable=True,
                        description="URL аватара сотрудника",
                    ),
                    "department": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Отдел, к которому относится сотрудник",
                    ),
                    "department_id": openapi.Schema(
                        type=openapi.TYPE_NUMBER,
                        description="Id отдела",
                    ),
                    "attendance": openapi.Schema(
                        type=openapi.TYPE_OBJECT,
                        additional_properties=openapi.Schema(
                            type=openapi.TYPE_OBJECT,
                            description="Данные о посещаемости",
                        ),
                    ),
                    "percent_for_period": openapi.Schema(
                        type=openapi.TYPE_NUMBER,
                        format=openapi.FORMAT_FLOAT,
                        description="Общий процент работы за указанный период",
                    ),
                    "salary": openapi.Schema(
                        type=openapi.TYPE_NUMBER,
                        format=openapi.FORMAT_FLOAT,
                        nullable=True,
                        description="Общая заработная плата сотрудника",
                    ),
                    "contract_type": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Тип контракта сотрудника",
                    ),
                },
            ),
        ),
        400: "Неверный запрос, дата начала не может быть позже даты окончания",
        404: "Сотрудник не найден",
    },
)
@api_view(["GET"])
@permission_classes([permissions.IsAuthenticatedOrAPIKey])
def staff_detail(request, staff_pin):
    """
    **Получить информацию о сотруднике**

    Данный метод возвращает подробную информацию о сотруднике, включая данные о посещаемости, заработной плате и типе контракта за указанный период.

    ### Args:
        - **request (HttpRequest)**: Запрос, содержащий параметры запроса.
        - **staff_pin (str)**: Уникальный идентификатор сотрудника (PIN).

    ### Returns:
        - **Response**: Ответ с данными сотрудника или сообщением об ошибке.

    ### Raises:
        - **ValueError**: Если start_date больше end_date.
    """

    logger.info(f"Request received for staff details with PIN {staff_pin}")

    staff = get_cache(
        f"staff_{staff_pin}", query=lambda: fetch_staff_data(staff_pin), timeout=1 * 10
    )

    if staff is None:
        logger.warning(f"Staff with PIN {staff_pin} not found")
        return Response(status=status.HTTP_404_NOT_FOUND)

    start_date, end_date = get_date_range(request)
    logger.debug(f"Retrieved date range: start_date={start_date}, end_date={end_date}")

    if start_date > end_date:
        logger.warning(
            f"Invalid date range: start_date {start_date} is greater than end_date {end_date}"
        )
        return Response(
            data={"error": "start_date cannot be greater than end_date"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    cache_key = f"staff_detail_{staff_pin}_{start_date}_{end_date}"
    logger.debug(f"Generated cache key: {cache_key}")

    data = get_cache(
        cache_key,
        query=lambda: get_staff_detail(staff, start_date, end_date),
        timeout=1 * 1 * 30,
    )

    logger.info(f"Returning staff details for PIN {staff_pin}")
    return Response(data, status=status.HTTP_200_OK)


def fetch_staff_data(staff_pin):
    """Получение данных о сотруднике из базы данных.
    Args:
        staff_pin (str): Уникальный идентификатор сотрудника (PIN).
    Returns:
        models.Staff: Объект сотрудника.
        None: Если сотрудник не найден.
    """
    try:
        return models.Staff.objects.get(pin=staff_pin)
    except models.Staff.DoesNotExist:
        return None


def get_date_range(request):
    """
    Получение диапазона дат из параметров запроса.

    Если даты не указаны, используется период последних 7 дней.

    Args:
        request (HttpRequest): Запрос с параметрами.

    Returns:
        tuple: Кортеж с датами начала и окончания периода (datetime.date).
    """
    end_date_str = request.query_params.get(
        "end_date", timezone.now().strftime("%Y-%m-%d")
    )
    start_date_str = request.query_params.get(
        "start_date", (timezone.now() - datetime.timedelta(days=7)).strftime("%Y-%m-%d")
    )

    end_date = datetime.datetime.strptime(end_date_str, "%Y-%m-%d").date()
    start_date = datetime.datetime.strptime(start_date_str, "%Y-%m-%d").date()

    return start_date, end_date


def get_staff_detail(staff, start_date, end_date):
    """
    Получение подробной информации о сотруднике за указанный период.

    Включает данные о посещаемости, процент присутствия, заработную плату и тип контракта.

    Args:
        staff (Staff): Объект сотрудника.
        start_date (datetime.date): Дата начала периода.
        end_date (datetime.date): Дата окончания периода.

    Returns:
        dict: Словарь с данными сотрудника.
    """
    logger.info(f"Получение деталей сотрудника {staff.name} (PIN: {staff.pin})")
    logger.debug(f"Запрошенный диапазон дат: {start_date} до {end_date}")

    attendance_qs = models.StaffAttendance.objects.filter(
        staff=staff,
        date_at__range=[
            start_date + datetime.timedelta(days=1),
            end_date + datetime.timedelta(days=1),
        ],
    )
    logger.debug(f"Получено записей о посещаемости: {attendance_qs.count()}")

    holidays = models.PublicHoliday.objects.filter(
        date__range=[start_date, end_date]
    ).values_list("date", "is_working_day")
    logger.debug(f"Государственные праздники в периоде: {len(holidays)}")

    holiday_dict = dict(holidays)
    attendance_data = {}
    total_minutes_for_period = 0
    total_days_with_data = 0
    percent_for_period = 0

    remote_work_qs = models.RemoteWork.objects.filter(staff=staff).filter(
        Q(permanent_remote=True) | Q(start_date__lte=end_date, end_date__gte=start_date)
    )
    logger.debug(f"Получено периодов дистанционной работы: {remote_work_qs.count()}")

    absent_reason_qs = models.AbsentReason.objects.filter(
        staff=staff, start_date__lte=end_date, end_date__gte=start_date
    )
    logger.debug(f"Получено причин отсутствия: {absent_reason_qs.count()}")

    dates = []

    if attendance_qs.exists():
        attendance_dates = [
            attendance.date_at - datetime.timedelta(days=1)
            for attendance in attendance_qs
        ]
        dates.extend(attendance_dates)
    else:
        attendance_dates = []

    if remote_work_qs.exists():
        remote_dates = []
        for remote_work in remote_work_qs:
            if remote_work.permanent_remote:
                remote_dates.extend(attendance_dates)
            else:
                remote_start = max(remote_work.start_date, start_date)
                remote_end = min(remote_work.end_date, end_date)
                remote_dates.extend(
                    [
                        remote_start + datetime.timedelta(days=x)
                        for x in range((remote_end - remote_start).days + 1)
                    ]
                )
        dates.extend(remote_dates)
    else:
        remote_dates = []

    if absent_reason_qs.exists():
        absent_dates = []
        for absent_reason in absent_reason_qs:
            absent_start = max(absent_reason.start_date, start_date)
            absent_end = min(absent_reason.end_date, end_date)
            absent_dates.extend(
                [
                    absent_start + datetime.timedelta(days=x)
                    for x in range((absent_end - absent_start).days + 1)
                ]
            )
        dates.extend(absent_dates)
    else:
        absent_dates = []

    if not dates:
        logger.warning(
            "Нет данных о посещаемости, дистанционной работе или причинах отсутствия за указанный период."
        )
        staff_detail = {
            "name": staff.name,
            "surname": staff.surname if staff.surname != "Нет фамилии" else "",
            "positions": [position.name for position in staff.positions.all()],
            "avatar": (
                staff.avatar.url if staff.avatar else "/media/images/no-avatar.png"
            ),
            "department": staff.department.name if staff.department else "N/A",
            "department_id": staff.department.id if staff.department else "N/A",
            "attendance": {},
            "percent_for_period": 0.0,
            "contract_type": None,
            "salary": None,
        }
        return staff_detail

    min_date = max(min(dates), start_date)
    max_date = min(max(dates), end_date)
    logger.debug(f"Фактический диапазон дат с данными: {min_date} до {max_date}")

    start_date = min_date
    end_date = max_date

    date_set = set(dates)
    logger.debug(f"Общее количество уникальных дат с данными: {len(date_set)}")

    num_days = len(date_set)
    cost_per_day = 100 / num_days
    logger.debug(
        f"Количество дней с данными: {num_days}, стоимость дня: {cost_per_day}"
    )

    average_attendance = get_average_attendance_for_period(staff, start_date, end_date)
    logger.debug(f"Средняя посещаемость за период: {average_attendance}")

    K_adj = 1.25
    penalty_rate = (100 / average_attendance) * K_adj
    logger.debug(f"Расчет штрафного коэффициента: {penalty_rate}")

    salary_qs = models.Salary.objects.filter(staff=staff).first()
    contract_type = salary_qs.contract_type if salary_qs else "full_time"
    total_minutes_expected_per_day = get_expected_minutes_per_day(contract_type)
    logger.debug(
        f"Тип контракта: {contract_type}, ожидаемые минуты в день: {total_minutes_expected_per_day}"
    )

    attendance_dict = {
        attendance.date_at - datetime.timedelta(days=1): attendance
        for attendance in attendance_qs
    }

    for event_date in sorted(date_set):
        logger.debug(f"Обработка даты: {event_date}")

        attendance = attendance_dict.get(event_date)
        (
            attendance_record,
            total_minutes_for_period,
            total_days_with_data,
            percent_for_period,
        ) = process_attendance(
            attendance,
            event_date,
            start_date,
            end_date,
            holiday_dict,
            total_minutes_expected_per_day,
            cost_per_day,
            penalty_rate,
            total_minutes_for_period,
            total_days_with_data,
            percent_for_period,
            remote_work_qs,
            absent_reason_qs,
        )

        if attendance_record:
            attendance_data[event_date.strftime("%d-%m-%Y")] = attendance_record
            logger.debug(
                f"Добавлена запись посещаемости для {event_date}: {attendance_record}"
            )

    if total_days_with_data > 0:
        percent_for_period /= total_days_with_data
        percent_for_period = max(percent_for_period, 0)
        logger.debug(f"Итоговый процент за период: {percent_for_period}")
    else:
        percent_for_period = 0.0
        logger.debug("Нет рабочих дней для расчета процента за период.")

    avatar_url = staff.avatar.url if staff.avatar else "/media/images/no-avatar.png"
    logger.debug(f"URL аватара: {avatar_url}")

    staff_detail = {
        "name": staff.name,
        "surname": staff.surname if staff.surname != "Нет фамилии" else "",
        "positions": [position.name for position in staff.positions.all()],
        "avatar": avatar_url,
        "department": staff.department.name if staff.department else "N/A",
        "department_id": staff.department.id if staff.department else "N/A",
        "attendance": attendance_data,
        "percent_for_period": round(percent_for_period, 2),
        "contract_type": salary_qs.contract_type if salary_qs else None,
        "salary": salary_qs.total_salary if salary_qs else None,
    }

    logger.info(
        f"Генерация деталей сотрудника завершена для {staff.name} (PIN: {staff.pin})"
    )
    return staff_detail


def get_average_attendance_for_period(staff, start_date, end_date):
    """
    Расчет среднего процента присутствия за предыдущий аналогичный период.

    Args:
        staff (Staff): Объект сотрудника.
        start_date (datetime.date): Дата начала текущего периода.
        end_date (datetime.date): Дата окончания текущего периода.

    Returns:
        float: Средний процент присутствия за предыдущий период.
    """
    logger.info(
        f"Calculating average attendance for staff {staff.name} (PIN: {staff.pin}) from {start_date} to {end_date}"
    )

    previous_start_date = start_date - datetime.timedelta(days=30)
    previous_end_date = end_date - datetime.timedelta(days=30)
    logger.debug(f"Previous period range: {previous_start_date} to {previous_end_date}")

    previous_attendance_qs = models.StaffAttendance.objects.filter(
        staff=staff, date_at__range=[previous_start_date, previous_end_date]
    )
    logger.debug(
        f"Retrieved {previous_attendance_qs.count()} attendance records for previous period"
    )

    if not previous_attendance_qs.exists():
        logger.warning(
            "No attendance records found for the previous period. Returning default average attendance of 85.0%"
        )
        return 85.0

    total_minutes = 0
    total_days = 0

    for attendance in previous_attendance_qs:
        first_in = attendance.first_in
        last_out = attendance.last_out

        if first_in and last_out:
            minutes_present = (last_out - first_in).total_seconds() / 60
            total_minutes += minutes_present
            total_days += 1
            logger.debug(
                f"Processed attendance for {attendance.date_at}: {minutes_present} minutes present"
            )

    if total_days == 0:
        logger.warning(
            "No complete attendance days found for the previous period. Returning default average attendance of 85.0%"
        )
        return 85.0

    average_attendance = (total_minutes / (total_days * 8 * 60)) * 100
    logger.info(
        f"Calculated average attendance for previous period: {average_attendance}%"
    )
    return average_attendance


def get_expected_minutes_per_day(contract_type):
    """
    Получение ожидаемого количества рабочих минут в день на основе типа контракта.

    Args:
        contract_type (str): Тип контракта сотрудника.

    Returns:
        int: Ожидаемые минуты в день.
    """
    if contract_type in ["part_time", "gph"]:
        return 4 * 60
    return 8 * 60


def process_attendance(
    attendance,
    event_date,
    start_date,
    end_date,
    holiday_dict,
    total_minutes_expected_per_day,
    cost_per_day,
    penalty_rate,
    total_minutes_for_period,
    total_days_with_data,
    percent_for_period,
    remote_work_qs,
    absent_reason_qs,
):
    """
    Обработка данных о посещаемости для конкретной даты с учетом новых требований.

    Args:
        attendance (StaffAttendance): Запись о посещаемости за дату, если есть.
        event_date (datetime.date): Дата, которую обрабатываем.
        start_date (datetime.date): Дата начала периода.
        end_date (datetime.date): Дата окончания периода.
        holiday_dict (dict): Словарь с информацией о праздничных днях.
        total_minutes_expected_per_day (int): Ожидаемое количество минут работы в день.
        cost_per_day (float): Стоимость одного дня в процентах.
        penalty_rate (float): Штрафной коэффициент за отсутствие.
        total_minutes_for_period (float): Общее количество минут за период.
        total_days_with_data (int): Общее количество дней с данными.
        percent_for_period (float): Процент рабочего времени за период.
        remote_work_qs (QuerySet): QuerySet с периодами дистанционной работы сотрудника.
        absent_reason_qs (QuerySet): QuerySet с причинами отсутствия сотрудника.

    Returns:
        tuple: Кортеж, содержащий:
            - attendance_record (dict): Обработанные данные о посещаемости за дату.
            - total_minutes_for_period (float): Обновленное общее количество минут за период.
            - total_days_with_data (int): Обновленное количество дней с данными.
            - percent_for_period (float): Обновленный процент рабочего времени за период.
    """
    logger.info(f"Обработка посещаемости за дату {event_date}")

    if not (start_date <= event_date <= end_date):
        logger.warning(
            f"Дата события {event_date} вне указанного диапазона {start_date} до {end_date}"
        )
        return None, total_minutes_for_period, total_days_with_data, percent_for_period

    is_off_day = check_off_day(event_date, holiday_dict)
    logger.debug(f"Является ли выходным днем: {is_off_day} для даты {event_date}")

    absent_reason = absent_reason_qs.filter(
        start_date__lte=event_date, end_date__gte=event_date
    ).first()

    first_in = attendance.first_in if attendance and attendance.first_in else None
    last_out = attendance.last_out if attendance and attendance.last_out else None

    if is_off_day:
        if first_in and last_out:
            total_minutes_worked = (last_out - first_in).total_seconds() / 60
            percent_day = (total_minutes_worked / total_minutes_expected_per_day) * 100
            logger.info(
                f"Сотрудник работал в выходной день {event_date}. Данные отображаются, но не влияют на расчеты."
            )
        else:
            total_minutes_worked = 0
            percent_day = 0
            logger.info(
                f"Выходной день {event_date} без данных о посещаемости. Пропускаем."
            )

        attendance_record = {
            "first_in": (
                first_in.astimezone(timezone.get_current_timezone())
                if first_in
                else None
            ),
            "last_out": (
                last_out.astimezone(timezone.get_current_timezone())
                if last_out
                else None
            ),
            "percent_day": round(percent_day, 2),
            "total_minutes": round(total_minutes_worked, 2),
            "is_weekend": True,
            "is_remote_work": False,
            "is_absent_approved": False,
            "absent_reason": None,
        }
        return (
            attendance_record,
            total_minutes_for_period,
            total_days_with_data,
            percent_for_period,
        )

    is_remote_work = remote_work_qs.filter(
        Q(permanent_remote=True)
        | Q(start_date__lte=event_date, end_date__gte=event_date)
    ).exists()

    is_absent_approved = False
    absent_reason_display = None

    if is_remote_work:
        percent_day = 100.0
        total_minutes_worked = total_minutes_expected_per_day
        total_minutes_for_period += total_minutes_worked
        total_days_with_data += 1
        percent_for_period += percent_day
        logger.info(f"{event_date} отмечен как день дистанционной работы.")
    elif absent_reason:
        is_absent_approved = absent_reason.approved
        absent_reason_display = absent_reason.get_reason_display()
        if is_absent_approved:
            logger.info(
                f"{event_date} утвержденная причина отсутствия: {absent_reason_display}."
            )
            attendance_record = {
                "first_in": (
                    first_in.astimezone(timezone.get_current_timezone())
                    if first_in
                    else None
                ),
                "last_out": (
                    last_out.astimezone(timezone.get_current_timezone())
                    if last_out
                    else None
                ),
                "percent_day": 0,
                "total_minutes": 0,
                "is_weekend": False,
                "is_remote_work": False,
                "is_absent_approved": True,
                "absent_reason": absent_reason_display,
            }
            return (
                attendance_record,
                total_minutes_for_period,
                total_days_with_data,
                percent_for_period,
            )
        else:
            percent_day = 0
            total_minutes_worked = 0
            total_days_with_data += 1
            penalty = penalty_rate * cost_per_day
            percent_for_period -= penalty
            logger.warning(
                f"{event_date} неутвержденная причина отсутствия: {absent_reason_display}. Применяется штраф {penalty}%."
            )
    else:
        if first_in and last_out:
            total_minutes_worked = (last_out - first_in).total_seconds() / 60
            percent_day = (total_minutes_worked / total_minutes_expected_per_day) * 100
            total_minutes_for_period += total_minutes_worked
            total_days_with_data += 1
            percent_for_period += percent_day
            logger.debug(
                f"Отработано минут: {total_minutes_worked}, Процент дня: {percent_day}"
            )
        else:
            percent_day = 0
            total_minutes_worked = 0
            total_days_with_data += 1
            penalty = penalty_rate * cost_per_day
            percent_for_period -= penalty
            logger.warning(
                f"Нет записей о посещаемости за дату {event_date}. Применяется штраф {penalty}%."
            )

    attendance_record = {
        "first_in": (
            first_in.astimezone(timezone.get_current_timezone()) if first_in else None
        ),
        "last_out": (
            last_out.astimezone(timezone.get_current_timezone()) if last_out else None
        ),
        "percent_day": round(percent_day, 2),
        "total_minutes": round(total_minutes_worked, 2),
        "is_weekend": is_off_day,
        "is_remote_work": is_remote_work,
        "is_absent_approved": is_absent_approved,
        "absent_reason": absent_reason_display,
    }
    logger.info(
        f"Обработана запись посещаемости за дату {event_date}: {attendance_record}"
    )

    return (
        attendance_record,
        total_minutes_for_period,
        total_days_with_data,
        percent_for_period,
    )


def check_off_day(event_date, holiday_dict):
    """
    Проверка, является ли дата выходным или праздничным днем.

    Args:
        event_date (datetime.date): Дата для проверки.
        holiday_dict (dict): Словарь с информацией о праздничных днях.

    Returns:
        bool: True, если день является выходным или праздничным, иначе False.
    """
    is_weekend = event_date.weekday() >= 5
    is_holiday = event_date in holiday_dict
    return (is_weekend and event_date not in holiday_dict) or (
        is_holiday and not holiday_dict[event_date]
    )


def update_percent_for_period(
    percent_for_period,
    percent_day,
    is_off_day,
    total_minutes_worked,
    cost_per_day,
    penalty_rate,
):
    """
    Обновление накопленного процента за период на основе ежедневной посещаемости.

    Args:
        percent_for_period (float): Текущий накопленный процент.
        percent_day (float): Процент присутствия за день.
        is_off_day (bool): Является ли день выходным.
        total_minutes_worked (float): Отработано минут за день.
        cost_per_day (float): Стоимость одного дня в процентах.
        penalty_rate (float): Штрафной коэффициент за отсутствие.

    Returns:
        float: Обновленный процент за период.
    """
    logger.debug(
        f"Updating percent for period. Initial: {percent_for_period}%, "
        f"Day percent: {percent_day}%, Is off day: {is_off_day}, "
        f"Total minutes worked: {total_minutes_worked}, "
        f"Cost per day: {cost_per_day}%, Penalty rate: {penalty_rate}%"
    )

    if is_off_day and total_minutes_worked > 0:
        percent_for_period += percent_day * 1.5
        logger.info(f"Off day with work. Increasing percent by {percent_day * 1.5}%.")
    elif not is_off_day and total_minutes_worked == 0:
        penalty = penalty_rate * cost_per_day
        percent_for_period -= penalty
        logger.warning(f"Workday with no work. Decreasing percent by {penalty}%.")
    else:
        percent_for_period += percent_day
        logger.info(f"Regular day. Adding {percent_day}% to the period percent.")

    logger.debug(f"Updated percent for period: {percent_for_period}%")
    return percent_for_period


@api_view(["GET"])
@permission_classes([permissions.IsAuthenticatedOrAPIKey])
def check_lesson_task_status(request, task_id):
    """
    Проверка статуса задачи и получение lesson_id.
    """
    try:
        task_result = AsyncResult(task_id)

        if task_result.state == "PENDING":
            return Response(
                {
                    "status": "Pending",
                    "message": "Задача в очереди, ожидайте завершения",
                },
                status=status.HTTP_202_ACCEPTED,
            )

        elif task_result.state == "SUCCESS":
            result = task_result.result
            return Response(
                {"status": "Success", "lesson_ids": result.get("success_records", [])},
                status=status.HTTP_200_OK,
            )

        elif task_result.state == "FAILURE":
            return Response(
                {"status": "Failure", "error": str(task_result.info)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        else:
            return Response(
                {
                    "status": task_result.state,
                    "message": "Задача в процессе выполнения",
                },
                status=status.HTTP_200_OK,
            )

    except Exception as e:
        logger.error(f"Ошибка при проверке задачи: {str(e)}")
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@swagger_auto_schema(
    method="post",
    operation_summary="Создание записей посещаемости занятий",
    operation_description=(
        "Создаёт новые записи посещаемости для сотрудников на занятия. "
        "Каждая запись должна содержать обязательные параметры: "
        "`staff_pin`, `subject_name`, `tutor_id`, `tutor`, `first_in`, `latitude`, `longitude`. "
        "`image` должен быть отправлен как отдельный файл в запросе."
    ),
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        required=["attendance_data", "image"],
        properties={
            "attendance_data": openapi.Schema(
                type=openapi.TYPE_ARRAY,
                items=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    required=[
                        "staff_pin",
                        "subject_name",
                        "tutor_id",
                        "tutor",
                        "first_in",
                        "latitude",
                        "longitude",
                    ],
                    properties={
                        "staff_pin": openapi.Schema(
                            type=openapi.TYPE_STRING,
                            description="PIN сотрудника",
                            example="s00260",
                        ),
                        "subject_name": openapi.Schema(
                            type=openapi.TYPE_STRING,
                            description="Название предмета",
                            example="Медицина",
                        ),
                        "tutor_id": openapi.Schema(
                            type=openapi.TYPE_INTEGER,
                            description="ID преподавателя",
                            example=1,
                        ),
                        "tutor": openapi.Schema(
                            type=openapi.TYPE_STRING,
                            description="ФИО преподавателя",
                            example="Иванов И.И.",
                        ),
                        "first_in": openapi.Schema(
                            type=openapi.TYPE_STRING,
                            format=openapi.FORMAT_DATETIME,
                            description="Время начала занятия в формате ISO 8601",
                            example="2024-10-06T14:24:24+05:00",
                        ),
                        "latitude": openapi.Schema(
                            type=openapi.TYPE_NUMBER,
                            format=openapi.FORMAT_FLOAT,
                            description="Широта места проведения",
                            example=43.207674,
                        ),
                        "longitude": openapi.Schema(
                            type=openapi.TYPE_NUMBER,
                            format=openapi.FORMAT_FLOAT,
                            description="Долгота места проведения",
                            example=76.851377,
                        ),
                    },
                ),
            ),
            "staff_image": openapi.Schema(
                type=openapi.TYPE_STRING,
                format=openapi.FORMAT_BINARY,
                description="Фотография сотрудника в бинарном формате.",
            ),
        },
    ),
    responses={
        202: openapi.Response(
            description="Задача по созданию записей принята в обработку",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "message": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Сообщение об успешном запуске задачи",
                    ),
                    "task_id": openapi.Schema(type=openapi.TYPE_STRING, description="ID задачи"),
                },
            ),
        ),
        400: openapi.Response(
            description="Неверные данные",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "error": openapi.Schema(
                        type=openapi.TYPE_STRING, description="Описание ошибки"
                    ),
                },
            ),
        ),
        500: openapi.Response(
            description="Ошибка сервера",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "error": openapi.Schema(
                        type=openapi.TYPE_STRING, description="Описание ошибки"
                    ),
                },
            ),
        ),
    },
)
@api_view(["POST"])
@permission_classes([permissions.IsAuthenticatedOrAPIKey])
def create_lesson_attendance(request):
    """
    Обрабатывает POST-запрос для создания записей посещаемости сотрудников на занятия.

    Функция принимает JSON-данные о посещаемости и файл с фотографией сотрудника, проверяет
    корректность данных и запускает асинхронную задачу Celery для сохранения данных и фотографии.

    Args:
        request (Request): Объект запроса, содержащий:
            - attendance_data (list): Список объектов с данными о посещаемости, каждый объект содержит:
                - staff_pin (str): PIN сотрудника.
                - subject_name (str): Название предмета.
                - tutor_id (int): ID преподавателя.
                - tutor (str): ФИО преподавателя.
                - first_in (str): Время начала занятия в формате ISO 8601.
                - latitude (float): Широта места проведения занятия.
                - longitude (float): Долгота места проведения занятия.
            - staff_image (File): Файл с фотографией сотрудника.

    Returns:
        Response:
            - 202 Accepted: Задача по созданию записей принята в обработку.
            - 400 Bad Request: Если переданы некорректные данные или отсутствуют обязательные параметры.
            - 500 Internal Server Error: В случае возникновения ошибки на сервере.

    Raises:
        Exception: Любые исключения логируются, и сервер возвращает ответ с кодом 500.
    """
    try:
        staff_image = request.FILES.get("image")
        attendance_data_raw = request.data.get("attendance_data")

        attendance_data = json.loads(attendance_data_raw)

        if not attendance_data or not staff_image:
            return Response(
                {"error": "Attendance data or staff image is missing"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        for record in attendance_data:
            required_fields = [
                "staff_pin",
                "subject_name",
                "tutor_id",
                "tutor",
                "first_in",
                "latitude",
                "longitude",
            ]
            if not all(record.get(field) for field in required_fields):
                return Response(
                    {"error": f"Missing required fields in record: {record}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        image_content = staff_image.read()

        task = tasks.process_lesson_attendance_batch.apply_async(
            args=[attendance_data, staff_image.name, image_content]
        )

        return Response(
            {"message": "Задача принята в обработку", "task_id": task.id},
            status=status.HTTP_202_ACCEPTED,
        )

    except Exception as e:
        logger.error(f"Ошибка при запуске задачи: {str(e)}")
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@swagger_auto_schema(
    method="put",
    operation_summary="Обновление записи посещаемости занятия",
    operation_description="Обновляет существующую запись посещаемости занятия по её ID. Параметр `last_out` обязателен, так как он указывает время окончания занятия. Параметры `first_in`, `latitude` и `longitude` могут быть обновлены опционально.",
    manual_parameters=[
        openapi.Parameter(
            "id",
            openapi.IN_PATH,
            description="ID записи для обновления",
            type=openapi.TYPE_INTEGER,
            required=True,
        ),
    ],
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        required=["last_out"],
        properties={
            "first_in": openapi.Schema(
                type=openapi.TYPE_STRING,
                format=openapi.FORMAT_DATETIME,
                description="Время начала занятия в формате ISO 8601 с часовым поясом. Опционально",
                example="2024-09-16T17:28:24+05:00",
            ),
            "last_out": openapi.Schema(
                type=openapi.TYPE_STRING,
                format=openapi.FORMAT_DATETIME,
                description="Время окончания занятия в формате ISO 8601 с часовым поясом. Обязательно",
                example="2024-09-16T18:28:24+05:00",
            ),
            "latitude": openapi.Schema(
                type=openapi.TYPE_NUMBER,
                format=openapi.FORMAT_FLOAT,
                description="Широта места проведения. Опционально",
                example=43.222,
            ),
            "longitude": openapi.Schema(
                type=openapi.TYPE_NUMBER,
                format=openapi.FORMAT_FLOAT,
                description="Долгота места проведения. Опционально",
                example=76.851,
            ),
        },
    ),
    responses={
        200: openapi.Response(
            description="Запись успешно обновлена",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "message": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Сообщение об успешном обновлении записи",
                    ),
                },
            ),
        ),
        400: openapi.Response(
            description="Неверные данные или отсутствует обязательный параметр `last_out`",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "error": openapi.Schema(type=openapi.TYPE_STRING),
                },
            ),
        ),
        404: openapi.Response(
            description="Запись с указанным ID не найдена",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "error": openapi.Schema(type=openapi.TYPE_STRING),
                },
            ),
        ),
        500: openapi.Response(
            description="Ошибка сервера",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "error": openapi.Schema(type=openapi.TYPE_STRING),
                },
            ),
        ),
    },
)
@api_view(["PUT"])
@permission_classes([permissions.IsAuthenticatedOrAPIKey])
def update_lesson_attendance(request, id):
    """
    Обновление записи посещаемости занятия.

    Args:
        id (int): ID записи для обновления.
        request (Request): HTTP запрос, содержащий данные для обновления записи.

    Ожидаемые параметры:
        last_out (str): Время окончания занятия в формате ISO 8601 с часовым поясом (обязательно).
        first_in (str): Время начала занятия в формате ISO 8601 с часовым поясом (опционально).
        latitude (float): Широта места проведения (опционально).
        longitude (float): Долгота места проведения (опционально).

    Returns:
        Response: Возвращает сообщение об успешном обновлении записи.
    """
    try:
        lesson_attendance = get_object_or_404(models.LessonAttendance, id=id)

        first_in = request.data.get("first_in", lesson_attendance.first_in)
        last_out = request.data.get("last_out")
        latitude = request.data.get("latitude", lesson_attendance.latitude)
        longitude = request.data.get("longitude", lesson_attendance.longitude)

        if not last_out:
            return Response(
                {"error": "'last_out' is required for updating."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        lesson_attendance.first_in = first_in
        lesson_attendance.last_out = last_out
        lesson_attendance.latitude = latitude
        lesson_attendance.longitude = longitude
        lesson_attendance.save()

        return Response(
            {
                "message": "LessonAttendance updated successfully",
                "lesson_id": lesson_attendance.id,
            },
            status=status.HTTP_200_OK,
        )

    except models.LessonAttendance.DoesNotExist:
        return Response(
            {"error": "LessonAttendance not found."}, status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


@swagger_auto_schema(
    method="get",
    operation_summary="Посещаемость сотрудников по отделу",
    operation_description="Получить данные о посещаемости сотрудников по ID подразделения",
    responses={
        200: openapi.Response(
            description="Успешный ответ",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "department_name": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Название подразделения",
                    ),
                    "attendance": openapi.Schema(
                        type=openapi.TYPE_ARRAY,
                        items=openapi.Schema(
                            type=openapi.TYPE_OBJECT,
                            properties={
                                "date": openapi.Schema(
                                    type=openapi.TYPE_ARRAY,
                                    items=openapi.Schema(
                                        type=openapi.TYPE_OBJECT,
                                        properties={
                                            "staff_fio": openapi.Schema(
                                                type=openapi.TYPE_STRING,
                                                description="ФИО сотрудника",
                                            ),
                                            "first_in": openapi.Schema(
                                                type=openapi.TYPE_STRING,
                                                format=openapi.FORMAT_DATETIME,
                                                description="Время первого входа сотрудника",
                                            ),
                                            "last_out": openapi.Schema(
                                                type=openapi.TYPE_STRING,
                                                format=openapi.FORMAT_DATETIME,
                                                description="Время последнего выхода сотрудника",
                                            ),
                                        },
                                    ),
                                    description="Список сотрудников и их посещаемость за дату",
                                )
                            },
                        ),
                        description="Список посещаемости по датам",
                    ),
                },
            ),
        ),
        400: openapi.Response(
            description="Ошибка в запросе",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "error": openapi.Schema(type=openapi.TYPE_STRING),
                },
            ),
        ),
        404: openapi.Response(
            description="Не найдено",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "error": openapi.Schema(type=openapi.TYPE_STRING),
                },
            ),
        ),
        500: openapi.Response(
            description="Ошибка сервера",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "error": openapi.Schema(type=openapi.TYPE_STRING),
                },
            ),
        ),
    },
    manual_parameters=[
        openapi.Parameter(
            "end_date",
            openapi.IN_QUERY,
            description="Конечная дата периода в формате YYYY-MM-DD",
            type=openapi.TYPE_STRING,
            required=True,
        ),
        openapi.Parameter(
            "start_date",
            openapi.IN_QUERY,
            description="Начальная дата периода в формате YYYY-MM-DD",
            type=openapi.TYPE_STRING,
            required=True,
        ),
    ],
)
@api_view(["GET"])
@permission_classes([AllowAny])
def staff_detail_by_department_id(request, department_id):
    """
    Получить данные о посещаемости сотрудников по ID подразделения.

    Этот эндпоинт возвращает данные о посещаемости сотрудников для указанного подразделения и его дочерних подразделений за указанный период.

    Параметры запроса:
    - end_date: Конечная дата периода в формате YYYY-MM-DD.
    - start_date: Начальная дата периода в формате YYYY-MM-DD.

    Возвращаемые данные:
    - department_name: Название подразделения.
    - attendance: Список посещаемости сотрудников, сгруппированных по датам.

    Пример ответа:
    {
        "department_name": "отдел цифровизации образовательных технологий",
        "attendance": [
            {
                "2024-07-01": [
                    {
                        "staff_fio": "Testov Test",
                        "first_in": "2024-07-01T00:00:00+05:00",
                        "last_out": "2024-07-01T23:59:59+05:00"
                    },
                    ...
                ]
            },
            ...
        ]
    }

    Возможные ошибки:
    - 400: Не указаны параметры начала или конца периода.
    - 404: Подразделение не найдено или данные о посещаемости не найдены.
    - 500: Внутренняя ошибка сервера.
    """
    logger.info(
        f"Request received for staff attendance by department ID {department_id}"
    )

    try:
        end_date_str = request.query_params.get("end_date")
        start_date_str = request.query_params.get("start_date")
        page = request.query_params.get("page", 1)

        if not end_date_str or not start_date_str:
            logger.warning("Missing start_date or end_date in request parameters")
            return Response(
                status=status.HTTP_400_BAD_REQUEST,
                data={"error": "Не указаны параметры начала или конца периода"},
            )

        start_date = datetime.datetime.strptime(start_date_str, "%Y-%m-%d")
        end_date = datetime.datetime.strptime(end_date_str, "%Y-%m-%d")
        logger.debug(f"Parsed date range: start_date={start_date}, end_date={end_date}")

        if start_date > end_date:
            logger.warning("Start date is greater than end date")
            return Response(
                status=status.HTTP_400_BAD_REQUEST,
                data={"error": "Дата начала не может быть больше даты конца"},
            )

        try:
            department = models.ChildDepartment.objects.get(id=department_id)
            logger.info(f"Department found: {department.name} (ID: {department_id})")
        except models.ChildDepartment.DoesNotExist:
            logger.warning(f"Department with ID {department_id} not found")
            return Response(
                status=status.HTTP_404_NOT_FOUND,
                data={"error": "Подразделение не найдено"},
            )

        def get_all_child_department_ids(department_id):
            child_ids = models.ChildDepartment.objects.filter(
                Q(parent_id=department_id) | Q(parent__parent_id=department_id)
            ).values_list("id", flat=True)
            return list(child_ids)

        department_ids = [department_id] + get_all_child_department_ids(department_id)
        logger.debug(f"Department IDs for attendance query: {department_ids}")

        cache_key = (
            f"staff_detail_{department_id}_{start_date_str}_{end_date_str}_page_{page}"
        )
        logger.debug(f"Generated cache key: {cache_key}")

        def query():
            logger.info("Querying staff attendance data")
            staff_attendance = (
                models.StaffAttendance.objects.filter(
                    staff__department_id__in=department_ids,
                    date_at__range=(start_date, end_date),
                )
                .select_related("staff")
                .order_by("date_at", "staff__surname", "staff__name")
            )

            paginator = StaffAttendancePagination()
            result_page = paginator.paginate_queryset(staff_attendance, request)
            serializer = serializers.StaffAttendanceByDateSerializer(
                result_page,
                many=True,
                context={"department_name": department.name},
            )

            logger.info("Staff attendance data query completed")
            return paginator.get_paginated_response(serializer.data).data

        cached_data = get_cache(cache_key, query=query, timeout=1 * 60 * 60)
        logger.info("Returning cached or queried data")
        return Response(cached_data)

    except Exception as e:
        logger.error(f"Server error while processing request: {str(e)}")
        return Response(
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            data={"error": str(e)},
        )


@swagger_auto_schema(
    method="post",
    operation_summary="Зарегистрировать нового пользователя (доступно только для администратора)",
    operation_description="Регистрирует нового пользователя в системе. Разрешено только для администратора.",
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        required=["username", "password"],
        properties={
            "username": openapi.Schema(
                type=openapi.TYPE_STRING,
                description="Желаемое имя для нового пользователя",
            ),
            "password": openapi.Schema(
                type=openapi.TYPE_STRING,
                description="Пароль для нового пользователя",
            ),
        },
    ),
    responses={
        201: openapi.Response(
            description="Created - Пользователь успешно зарегистрирован",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "message": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Сообщение о результате регистрации",
                    )
                },
            ),
        ),
        400: openapi.Response(
            description="Bad Request - Ошибка в запросе",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "message": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Описание ошибки запроса",
                    ),
                },
            ),
        ),
    },
)
@api_view(http_method_names=["POST"])
@permission_classes([IsAdminUser])
def user_register(request):
    """
    Регистрирует нового пользователя в системе. Разрешено только для администратора.

    Этот view ожидает запрос POST, содержащий в теле запроса следующие данные:
    - username (str): Желаемое имя для нового пользователя.
    - password (str): Пароль для нового пользователя.

    Возвращаемые данные:
    - status (int): HTTP статус код:
        - 201 Created: Если пользователь успешно зарегистрирован.
        - 400 Bad Request: Если имя пользователя или пароль отсутствуют, не соответствуют требованиям или имя пользователя уже занято.
    - message (str): Сообщение о результате регистрации.

    Возможные ошибки:
    - 400 Bad Request: Если имя пользователя или пароль отсутствуют, пароль не соответствует требованиям или имя пользователя уже занято.

    Исключения:
    - Стандартные исключения Django, если во время создания или сохранения пользователя возникают какие-либо ошибки.
    """

    logger.info("Received request to register a new user")

    username = request.data.get("username", None)
    password = request.data.get("password", None)

    if not username or not password:
        logger.warning("Username or password not provided")
        return Response(
            status=status.HTTP_400_BAD_REQUEST,
            data={"message": "Требуется юзернейм и пароль"},
        )

    if not utils.password_check(password):
        logger.warning("Password does not meet the requirements")
        return Response(
            status=status.HTTP_400_BAD_REQUEST,
            data={"message": "Пароль не прошел требования"},
        )

    try:
        user, created = User.objects.get_or_create(username=username)
        if not created:
            logger.warning(f"Username '{username}' is already taken")
            return Response(
                status=status.HTTP_400_BAD_REQUEST,
                data={"message": "Данный username уже занят"},
            )

        user.set_password(password)
        user.save()
        logger.info(f"User '{username}' successfully created")
        return Response(
            status=status.HTTP_201_CREATED,
            data={"message": "пользователь успешно создан"},
        )
    except Exception as e:
        logger.error(f"Error occurred while creating user '{username}': {str(e)}")
        return Response(
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            data={"message": str(e)},
        )


@swagger_auto_schema(
    method="get",
    operation_summary="Получить профиль пользователя",
    operation_description="Получить профиль текущего аутентифицированного пользователя.",
    responses={
        200: openapi.Response(
            description="Данные профиля пользователя",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "is_banned": openapi.Schema(
                        type=openapi.TYPE_BOOLEAN,
                        description="Забанен ли пользователь",
                    ),
                    "user": openapi.Schema(
                        type=openapi.TYPE_OBJECT,
                        properties={
                            "username": openapi.Schema(
                                type=openapi.TYPE_STRING,
                                description="Имя пользователя",
                            ),
                            "email": openapi.Schema(
                                type=openapi.TYPE_STRING,
                                format=openapi.FORMAT_EMAIL,
                                description="Электронная почта",
                            ),
                            "first_name": openapi.Schema(
                                type=openapi.TYPE_STRING,
                                description="Имя",
                            ),
                            "last_name": openapi.Schema(
                                type=openapi.TYPE_STRING,
                                description="Фамилия",
                            ),
                            "is_staff": openapi.Schema(
                                type=openapi.TYPE_BOOLEAN,
                                description="Является ли пользователь сотрудником",
                            ),
                            "date_joined": openapi.Schema(
                                type=openapi.TYPE_STRING,
                                format=openapi.FORMAT_DATETIME,
                                description="Дата регистрации",
                            ),
                            "last_login": openapi.Schema(
                                type=openapi.TYPE_STRING,
                                format=openapi.FORMAT_DATETIME,
                                description="Дата последнего входа",
                            ),
                            "phonenumber": openapi.Schema(
                                type=openapi.TYPE_STRING,
                                description="Номер телефона",
                            ),
                            "last_login_ip": openapi.Schema(
                                type=openapi.TYPE_STRING,
                                format=openapi.FORMAT_IPV4,
                                description="Последний IP-адрес входа",
                            ),
                        },
                    ),
                },
            ),
        ),
        401: openapi.Response(
            description="Unauthorized: Если пользователь не аутентифицирован.",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "detail": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Сообщение об ошибке аутентификации",
                    )
                },
            ),
        ),
    },
)
@swagger_auto_schema(
    method="put",
    operation_summary="Обновить профиль пользователя",
    operation_description="Обновить профиль текущего аутентифицированного пользователя. Можно обновлять одно или несколько полей: first_name, last_name, password, email, phonenumber. Если поле не отправлено, оно останется неизменным.",
    manual_parameters=[token_param_config],
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            "first_name": openapi.Schema(
                type=openapi.TYPE_STRING,
                description="Имя пользователя (опционально)",
            ),
            "last_name": openapi.Schema(
                type=openapi.TYPE_STRING,
                description="Фамилия пользователя (опционально)",
            ),
            "password": openapi.Schema(
                type=openapi.TYPE_STRING,
                description="Пароль пользователя (опционально)",
            ),
            "email": openapi.Schema(
                type=openapi.TYPE_STRING,
                format=openapi.FORMAT_EMAIL,
                description="Электронная почта пользователя (опционально)",
            ),
            "phonenumber": openapi.Schema(
                type=openapi.TYPE_STRING,
                description="Номер телефона пользователя (опционально)",
            ),
        },
    ),
    responses={
        200: openapi.Response(
            description="Профиль пользователя обновлен",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "user": openapi.Schema(
                        type=openapi.TYPE_OBJECT,
                        properties={
                            "username": openapi.Schema(
                                type=openapi.TYPE_STRING,
                                description="Имя пользователя",
                            ),
                            "email": openapi.Schema(
                                type=openapi.TYPE_STRING,
                                format=openapi.FORMAT_EMAIL,
                                description="Электронная почта",
                            ),
                            "first_name": openapi.Schema(
                                type=openapi.TYPE_STRING,
                                description="Имя",
                            ),
                            "last_name": openapi.Schema(
                                type=openapi.TYPE_STRING,
                                description="Фамилия",
                            ),
                            "phonenumber": openapi.Schema(
                                type=openapi.TYPE_STRING,
                                description="Номер телефона",
                            ),
                            "is_banned": openapi.Schema(
                                type=openapi.TYPE_BOOLEAN,
                                description="Забанен ли пользователь",
                            ),
                            "last_login_ip": openapi.Schema(
                                type=openapi.TYPE_STRING,
                                format=openapi.FORMAT_IPV4,
                                description="Последний IP-адрес входа",
                            ),
                        },
                    ),
                },
            ),
        ),
        400: openapi.Response(
            description="Bad Request: Неверные данные",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "detail": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Сообщение об ошибке",
                    )
                },
            ),
        ),
        401: openapi.Response(
            description="Unauthorized: Если пользователь не аутентифицирован.",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "detail": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Сообщение об ошибке аутентификации",
                    )
                },
            ),
        ),
    },
)
@api_view(["GET", "PUT"])
@permission_classes([IsAuthenticated])
def user_profile_detail(request):
    """
    Получить или обновить профиль текущего аутентифицированного пользователя.

    Метод GET возвращает данные профиля пользователя.
    Метод PUT позволяет обновлять поля last_name, first_name, password, email и phonenumber.

    Аргументы:
    - request: объект запроса.

    Возвращаемые данные:
    - GET: Ответ с данными профиля пользователя.
    - PUT: Ответ с обновленными данными профиля пользователя или сообщением об ошибке.

    Возможные ошибки:
    - 401: Если пользователь не аутентифицирован.
    - 404: Если профиль пользователя не найден.
    """

    def get_client_ip(request):
        x_forwarded_for = request.META.get("HTTP_X_FORWARDED_FOR")
        if x_forwarded_for:
            ip = x_forwarded_for.split(",")[0]
        else:
            ip = request.META.get("REMOTE_ADDR")
        return ip

    try:
        user_profile = models.UserProfile.objects.get(user=request.user)

        if request.method == "GET":
            serializer = serializers.UserProfileSerializer(user_profile)
            return Response(serializer.data)

        elif request.method == "PUT":
            data = request.data
            user = request.user
            update_user_fields = []
            update_profile_fields = []

            if "first_name" in data and data["first_name"] != user.first_name:
                user.first_name = data["first_name"]
                update_user_fields.append("first_name")

            if "last_name" in data and data["last_name"] != user.last_name:
                user.last_name = data["last_name"]
                update_user_fields.append("last_name")

            if "email" in data and data["email"] != user.email:
                user.email = data["email"]
                update_user_fields.append("email")

            if "password" in data:
                user.set_password(data["password"])
                update_user_fields.append("password")

            if (
                "phonenumber" in data
                and data["phonenumber"] != user_profile.phonenumber
            ):
                user_profile.phonenumber = data["phonenumber"]
                update_profile_fields.append("phonenumber")

            client_ip = get_client_ip(request)
            if client_ip != user_profile.last_login_ip:
                user_profile.last_login_ip = client_ip
                update_profile_fields.append("last_login_ip")

            if update_user_fields:
                user.save(update_fields=update_user_fields)

            if update_profile_fields:
                user_profile.save(update_fields=update_profile_fields)

            logger.info(
                f"User {request.user.username} updated their profile. Updated fields: {update_user_fields + update_profile_fields}"
            )

            serializer = serializers.UserProfileSerializer(user_profile)
            return Response(serializer.data, status=status.HTTP_200_OK)

    except models.UserProfile.DoesNotExist:
        logger.error(f"UserProfile does not exist for user {request.user.username}")
        return Response(
            status=status.HTTP_404_NOT_FOUND,
            data={"message": "Профиль пользователя не найден"},
        )
    except Exception as e:
        logger.error(
            f"Error while updating profile for user {request.user.username}: {str(e)}"
        )
        return Response(
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            data={
                "message": "Произошла ошибка при обновлении профиля",
                "error": str(e),
            },
        )


def logout_view(request):
    logout(request)
    return redirect("login_view")


def login_view(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect("uploadFile")

    return render(request, "login.html", context={})


@swagger_auto_schema(
    method="get",
    operation_summary="Запрос на получение данных с Внешнего сервера",
    operation_description="Запрос на получение данных о посещаемости. Требует передачи заголовка X-API-KEY для аутентификации.",
    manual_parameters=[
        openapi.Parameter(
            name="X-API-KEY",
            in_=openapi.IN_HEADER,
            type=openapi.TYPE_STRING,
            required=True,
            description="API ключ для аутентификации запроса.",
        ),
    ],
    responses={
        200: openapi.Response(
            description="Запуск процесса получения данных.",
            schema=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    "message": openapi.Schema(
                        type=openapi.TYPE_STRING,
                        description="Статус сообщения.",
                    ),
                },
            ),
        ),
        403: "Forbidden: Если доступ запрещен или отсутствует API ключ.",
        500: "Internal Server Error: В случае ошибки сервера.",
    },
)
@api_view(http_method_names=["GET"])
@permission_classes([permissions.IsAuthenticatedOrAPIKey])
def fetch_data_view(request):
    """
    Запрос на получение данных о посещаемости. Требует передачи заголовка X-API-KEY для аутентификации.

    Args:
    запрос: объект запроса.

    Returns:
    Ответ: статус сообщения.

    Raises:
    Http403: Если доступ запрещен или отсутствует API ключ.
    Http500: В случае ошибки сервера.
    """
    function_name = "fetch_data_view"
    start_time = time.perf_counter()
    logger.info(f"{function_name}: Request received")

    try:
        api_key = request.headers.get("X-API-KEY")
        if api_key is None:
            logger.warning(f"{function_name}: API key not provided in request headers")
            return Response(
                status=status.HTTP_403_FORBIDDEN,
                data={"error": "Доступ запрещен. Не указан API ключ."},
            )

        try:
            key_obj = models.APIKey.objects.get(key=api_key)
            if not key_obj.is_active:
                logger.warning(f"{function_name}: API key {api_key} is inactive")
                return Response(
                    status=status.HTTP_403_FORBIDDEN,
                    data={"error": "Доступ запрещен. Недействительный API ключ."},
                )
        except models.APIKey.DoesNotExist:
            logger.warning(f"{function_name}: API key {api_key} does not exist")
            return Response(
                status=status.HTTP_403_FORBIDDEN,
                data={"error": "Доступ запрещен. Недействительный API ключ."},
            )

        utils.get_all_attendance()
        logger.info(f"{function_name}: Attendance data fetched successfully")
        return Response(status=status.HTTP_200_OK, data={"message": "Done"})

    except Exception as e:
        logger.error(
            f"{function_name}: Error occurred while fetching attendance data: {str(e)}"
        )
        return Response(
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            data={"error": str(e)},
        )
    finally:
        end_time = time.perf_counter()
        duration_seconds = end_time - start_time
        duration_human_readable = utils.format_duration(duration_seconds)
        logger.info(
            f"{function_name} completed in {duration_human_readable} (({duration_seconds:.2f} seconds))"
        )


@swagger_auto_schema(
    method="get",
    operation_summary="Получение данных о посещаемости в формате Excel",
    operation_description="Получение данных о посещаемости с внешнего сервера и создание Excel файла. Требуется аутентификация.",
    manual_parameters=[
        openapi.Parameter(
            name="X-API-KEY",
            in_=openapi.IN_HEADER,
            type=openapi.TYPE_STRING,
            required=True,
            description="API ключ для аутентификации запроса.",
        ),
        openapi.Parameter(
            name="endDate",
            in_=openapi.IN_QUERY,
            type=openapi.TYPE_STRING,
            required=True,
            description="Конечная дата для данных о посещаемости в формате YYYY-MM-DD.",
        ),
        openapi.Parameter(
            name="startDate",
            in_=openapi.IN_QUERY,
            type=openapi.TYPE_STRING,
            required=True,
            description="Начальная дата для данных о посещаемости в формате YYYY-MM-DD.",
        ),
    ],
    responses={
        200: openapi.Response(
            description="Excel файл с данными о посещаемости.",
            schema=openapi.Schema(
                type=openapi.TYPE_FILE,
                description="Созданный Excel файл, содержащий данные о посещаемости.",
            ),
        ),
        400: openapi.Response(
            description="Bad Request: отсутствует начальная или конечная дата."
        ),
        403: openapi.Response(
            description="Forbidden: если доступ запрещен или отсутствует API ключ."
        ),
        500: openapi.Response(
            description="Internal server error: если сервер столкнулся с ошибкой."
        ),
    },
)
@api_view(http_method_names=["GET"])
@permission_classes([permissions.IsAuthenticatedOrAPIKey])
@utils.add_api_key_header
def sent_excel(request, department_id):
    """
    Получение данных о посещаемости в формате Excel.

    Получение данных о посещаемости с внешнего сервера на основе предоставленного ID отдела,
    начальной и конечной дат. Создание и возврат Excel файла, содержащего данные о посещаемости.

    Аргументы:
        request (HttpRequest): Объект HTTP запроса.
        department_id (int): ID отдела, для которого запрашиваются данные о посещаемости.

    Параметры запроса:
        - startDate (str): Начальная дата для данных о посещаемости в формате YYYY-MM-DD.
        - endDate (str): Конечная дата для данных о посещаемости в формате YYYY-MM-DD.

    Возвращает:
        HttpResponse: HTTP ответ с созданным Excel файлом или сообщением об ошибке.

    Исключения:
        ValueError: Если начальная или конечная дата отсутствует в параметрах запроса.
        ConnectionError: Если возникла проблема с получением данных с внешнего сервера.
    """
    logger.info(
        f"Request received to generate Excel file for department ID {department_id}"
    )

    end_date = request.query_params.get("endDate", None)
    start_date = request.query_params.get("startDate", None)

    if not all([end_date, start_date]):
        logger.warning(
            f"Missing startDate or endDate in request parameters for department ID {department_id}"
        )
        return Response(
            {"error": "Missing startDate or endDate"},
            status=status.HTTP_400_BAD_REQUEST,
        )
    try:
        end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d").date()
        start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d").date()
    except ValueError as e:
        logger.error(f"Invalid date format for department ID {department_id}: {e}")
        return Response(
            {"error": f"Invalid date format: {e}"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    logger.debug(
        f"Parsed dates for department ID {department_id}: start_date={start_date}, end_date={end_date}"
    )

    end_date += datetime.timedelta(days=1)
    start_date += datetime.timedelta(days=1)

    main_ip = settings.MAIN_IP
    rows = []
    page = 1

    while True:
        cache_key = f"{main_ip}_api_department_stats_{department_id}_{start_date}_{end_date}_page_{page}"
        url = f"{main_ip}/api/department/stats/{department_id}/?end_date={end_date}&start_date={start_date}&page={page}"

        logger.debug(f"Fetching data for department ID {department_id} from URL: {url}")

        data = get_cache(
            cache_key, query=lambda: utils.fetch_data(url), timeout=1 * 60 * 60
        )

        if "results" not in data or not data["results"]:
            logger.info(
                f"No more results found for department ID {department_id}, stopping data fetch."
            )
            break

        with ThreadPoolExecutor(max_workers=3) as executor:
            future = executor.submit(utils.parse_attendance_data, data["results"])
            rows += future.result()
            logger.debug(
                f"Processed page {page} for department ID {department_id}, rows collected: {len(rows)}"
            )

        if not data.get("next"):
            logger.info(
                f"All pages processed for department ID {department_id}, total rows collected: {len(rows)}"
            )
            break
        page += 1

    if rows:
        logger.info(
            f"Creating Excel file for department ID {department_id} with {len(rows)} rows"
        )
        df_pivot_sorted = utils.create_dataframe(rows)
        wb = utils.save_to_excel(df_pivot_sorted)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f"attachment; filename=Посещаемость_{department_id}.xlsx"
        )

        with NamedTemporaryFile(delete=False) as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            response.write(tmp.read())

        logger.info(
            f"Excel file created successfully for department ID {department_id}"
        )
        return response
    else:
        logger.error(
            f"Failed to generate Excel file for department ID {department_id}, no data rows found"
        )
        return Response(
            {"error": "Failed to generate Excel file"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


class UploadFileView(View):
    """
    Класс представления для обработки действий по загрузке файлов.

    Отображает форму загрузки файла (upload_file.html)
    и обрабатывает POST-запросы для импорта данных из файла.
    """

    template_name = "upload_file.html"

    def get(self, request, *args, **kwargs):
        """
        Обрабатывает GET-запросы.

        Args:
            request (HttpRequest): Объект запроса.
            *args: Дополнительные позиционные аргументы.
            **kwargs: Дополнительные именованные аргументы.

        Returns:
            HttpResponse: Отрисовывает шаблон upload_file.html
            с контекстом, содержащим список всех категорий файлов (categories) и список родительских отделов (parent_departments).
        """
        logger.info("GET request received for file upload view")
        categories = models.FileCategory.objects.all()
        parent_departments = models.ParentDepartment.objects.exclude(id=1)
        context = {"categories": categories, "parent_departments": parent_departments}
        logger.debug(f"Rendering template with context: {context}")
        return render(request, self.template_name, context=context)

    def post(self, request, *args, **kwargs):
        """
        Обрабатывает POST-запросы.

        Args:
            request (HttpRequest): Объект запроса.
            *args: Дополнительные позиционные аргументы.
            **kwargs: Дополнительные именованные аргументы.

        Returns:
            HttpResponse: Возвращает редирект на страницу загрузки файла или
            рендеринг upload_file.html с соответствующим контекстом.

        Raises:
            Exception: Если произошла ошибка при обработке файла.
        """
        logger.info("POST request received for file upload view")
        file_path = request.FILES.get("file")
        category_slug = request.POST.get("category")
        parent_department_id = request.POST.get("parent_department")

        if file_path and category_slug:
            logger.debug(f"File received: {file_path.name}, Category: {category_slug}")
            try:
                if file_path.name.endswith(".xlsx"):
                    logger.info("Processing Excel file")
                    rows = self.handle_excel(file_path)
                    if category_slug == "delete_staff":
                        logger.info("Deleting staff based on Excel data")
                        self.delete_staff(request, rows, parent_department_id)
                    elif category_slug == "staff":
                        logger.info("Processing staff data from Excel")
                        self.process_staff(request, rows)
                    elif category_slug == "departments":
                        logger.info("Processing departments data from Excel")
                        self.process_departments(request, rows)
                    messages.success(
                        request, "Файл успешно обработан и данные обновлены."
                    )
                elif file_path.name.endswith(".zip") and category_slug == "photo":
                    logger.info("Processing ZIP file for photos")
                    self.handle_zip(request, file_path)
                    messages.success(request, "Фото успешно загружены.")
                else:
                    logger.warning("Invalid file format or category")
                    messages.error(request, "Неверный формат файла или категория.")
                    return render(request, self.template_name)

                return redirect("uploadFile")
            except Exception as error:
                logger.error(f"Error processing file: {str(error)}")
                messages.error(request, f"Ошибка при обработке файла: {str(error)}")
        else:
            logger.warning("File or category missing in the POST request")
            messages.error(
                request,
                "Проверьте правильность заполненных данных или неверный формат файла.",
            )

        return render(request, self.template_name)

    @transaction.atomic
    def handle_excel(self, file_path):
        """
        Обрабатывает загрузку и импорт данных из файла Excel.

        Args:
            file_path (File): Путь к загруженному файлу.
            category_slug (str): Категория файла для обработки.

        Raises:
            Exception: Если произошла ошибка при обработке файла Excel.
        """
        logger.info("Handling Excel file")
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            ws.delete_rows(1, 2)
            rows = list(ws.iter_rows())
            logger.debug(f"Rows before sorting: {[row[0].value for row in rows]}")

            rows.sort(
                key=lambda row: (
                    not str(row[0].value).isdigit(),
                    str(row[0].value).zfill(10),
                ),
                reverse=False,
            )
            logger.debug(f"Rows after sorting: {[row[0].value for row in rows]}")
            logger.debug(f"Excel file processed, number of rows: {len(rows)}")
            return rows
        except Exception as e:
            logger.error(f"Error processing Excel file: {str(e)}")
            raise

    def delete_staff(self, request, rows, parent_department_id):
        """
        Удаляет сотрудников дочерних отделов, отсутствующих в переданном списке PIN-кодов.

        Метод получает родительский отдел по `parent_department_id` и находит все связанные
        дочерние отделы. Затем проверяет, какие PIN-коды сотрудников из базы данных
        отсутствуют в списке, переданном в `rows`, и удаляет таких сотрудников.

        Args:
            request: HTTP-запрос для отправки сообщений об успешном или неудачном удалении.
            rows: Список строк с PIN-кодами сотрудников, которых нужно оставить.
            parent_department_id: ID родительского отдела для поиска связанных дочерних отделов.

        Exceptions:
            ValueError: Если не передан `parent_department_id` или не найдены дочерние отделы.
            models.ParentDepartment.DoesNotExist: Если родительский отдел с данным ID не найден.
            Exception: Любая другая ошибка, возникшая при удалении сотрудников.
        """
        logger.info(f"Deleting staff for parent department ID: {parent_department_id}")
        try:
            if not parent_department_id:
                raise ValueError("ID родительского отдела не был передан.")

            parent_department = models.ParentDepartment.objects.get(
                id=parent_department_id
            )

            child_departments = models.ChildDepartment.objects.filter(
                parent__name=parent_department.name
            )
            if not child_departments.exists():
                raise ValueError(
                    f"Для родительского отдела {parent_department.name} не найдены дочерние отделы."
                )

            pin_list_from_file = [row[0].value for row in rows if row[0].value]

            staff_in_db = models.Staff.objects.filter(department__in=child_departments)

            staff_to_delete = staff_in_db.exclude(pin__in=pin_list_from_file)

            deleted_count, _ = staff_to_delete.delete()
            logger.info(f"Deleted {deleted_count} staff members")
            messages.success(
                request, f"Успешно удалено {deleted_count} сотрудника(ов)."
            )

        except models.ParentDepartment.DoesNotExist:
            error_message = f"Родительский отдел с ID {parent_department_id} не найден."
            logger.error(error_message)
            messages.error(request, error_message)
        except ValueError as ve:
            logger.warning(f"ValueError during staff deletion: {str(ve)}")
            messages.error(request, str(ve))
        except Exception as e:
            logger.error(f"Unexpected error during staff deletion: {str(e)}")
            messages.error(
                request, f"Произошла ошибка при удалении сотрудников: {str(e)}"
            )

    def process_departments(self, request, rows):
        """
        Обрабатывает данные для категории "departments" из Excel файла.
        Args:
            rows (list): Список строк из файла Excel.
        Raises:
            Exception: Если произошла ошибка при обработке строки.
        """
        logger.info("Processing departments data")

        created_parent_departments = []
        created_child_departments = []

        try:
            for row in rows:
                parent_department_id_value = row[2].value
                parent_department_name = row[3].value
                child_department_name = row[1].value
                child_department_id_value = row[0].value

                parent_department_id = (
                    utils.normalize_id(str(parent_department_id_value).strip())
                    if parent_department_id_value
                    else None
                )
                child_department_id = (
                    utils.normalize_id(str(child_department_id_value).strip())
                    if child_department_id_value
                    else None
                )

                if not parent_department_id or not child_department_id:
                    logger.debug("Skipping row due to missing or invalid ID")
                    continue

                if parent_department_name:
                    parent_department, parent_created = (
                        models.ParentDepartment.objects.get_or_create(
                            id=parent_department_id,
                            defaults={"name": parent_department_name},
                        )
                    )
                    if parent_created:
                        created_parent_departments.append(parent_department_name)
                        logger.info(
                            f"Created new parent department: {parent_department_name}"
                        )

                    parent_department_as_child, child_created = (
                        models.ChildDepartment.objects.get_or_create(
                            id=parent_department.id,
                            defaults={"name": parent_department.name, "parent": None},
                        )
                    )
                else:
                    parent_department_as_child = models.ChildDepartment.objects.get(
                        id="1"
                    )

                child_department, child_created = (
                    models.ChildDepartment.objects.get_or_create(
                        id=child_department_id,
                        defaults={
                            "name": child_department_name,
                            "parent": parent_department_as_child,
                        },
                    )
                )
                if child_created:
                    created_child_departments.append(child_department_name)
                    logger.info(
                        f"Created new child department: {child_department_name}"
                    )

            if created_parent_departments or created_child_departments:
                messages.success(
                    request,
                    f"Создано родительских отделов: {len(created_parent_departments)}, "
                    f"дочерних отделов: {len(created_child_departments)}.",
                )

        except Exception as error:
            logger.error(f"Error processing departments: {str(error)}")
            messages.error(request, f"Ошибка при обработке отдела: {str(error)}")

    def process_staff(self, request, rows):
        """
        Обрабатывает данные для категории "staff" из Excel файла.
        except Exception as error:
            messages.error(request, f"Ошибка при обработке отдела: {str(error)}")

        Args:
            rows (list): Список строк из файла Excel.

        Raises:
            Exception: Если произошла ошибка при обработке строки.
        """
        logger.info("Processing staff data")
        staff_instances = []
        departments_cache = {}

        try:
            for row in rows:
                pin = row[0].value
                name = row[1].value
                surname = row[2].value or "Нет фамилии"
                department_id = str(row[3].value) if row[3].value else None
                position_name = (
                    row[5].value or "Сотрудник" if len(row) > 5 else "Сотрудник"
                )

                position, _ = models.Position.objects.get_or_create(name=position_name)

                if department_id:
                    if department_id in departments_cache:
                        department = departments_cache[department_id]
                    else:
                        try:
                            department = models.ChildDepartment.objects.get(
                                id=department_id
                            )
                            departments_cache[department_id] = department
                        except models.ChildDepartment.DoesNotExist:
                            department = None
                else:
                    department = None

                staff_instance = models.Staff(
                    pin=pin,
                    name=name,
                    surname=surname,
                    department=department,
                )

                staff_instances.append((staff_instance, position))

            pin_list = [staff[0].pin for staff in staff_instances]
            existing_staff = models.Staff.objects.filter(pin__in=pin_list)
            existing_staff_dict = {staff.pin: staff for staff in existing_staff}

            staff_to_create = []
            staff_to_update = []

            for staff_instance, position in staff_instances:
                if staff_instance.pin in existing_staff_dict:
                    existing = existing_staff_dict[staff_instance.pin]
                    if staff_instance.name and staff_instance.name != existing.name:
                        existing.name = staff_instance.name
                    if (
                        staff_instance.surname
                        and staff_instance.surname != existing.surname
                    ):
                        existing.surname = staff_instance.surname
                    if (
                        staff_instance.department
                        and staff_instance.department != existing.department
                    ):
                        existing.department = staff_instance.department
                    if position.name and position.name != "Сотрудник":
                        if not existing.positions.filter(name=position.name).exists():
                            existing.positions.add(position)
                    staff_to_update.append(existing)
                else:
                    try:
                        staff_instance.save()
                        staff_instance.positions.add(position)
                        staff_to_create.append(staff_instance)
                    except IntegrityError:
                        logger.warning(
                            f"Duplicate entry for pin {staff_instance.pin}, skipping."
                        )
                        continue

            for staff in staff_to_update:
                staff.save()

            logger.info(f"Updated {len(staff_to_update)} staff members")
            logger.info(f"Created {len(staff_to_create)} new staff members")
            messages.success(
                request, f"Успешно обновлено {len(staff_to_update)} сотрудников."
            )
            messages.success(
                request, f"Успешно добавлено {len(staff_to_create)} новых сотрудников."
            )

        except Exception as e:
            logger.error(f"Error processing staff data: {str(e)}")
            messages.error(request, f"Ошибка при обработке сотрудников: {str(e)}")

    def handle_zip(self, request, file_path):
        """
        Обрабатывает загрузку и импорт данных из ZIP архива для Staff.
        Args:
            file_path (File): Путь к загруженному файлу.

        Raises:
            Exception: Если произошла ошибка при обработке ZIP файла.
        """
        logger.info("Processing ZIP file for staff photos")
        try:
            with zipfile.ZipFile(file_path, "r") as zip_file:
                zip_file.extractall("/tmp")
                for filename in zip_file.namelist():
                    pin = os.path.splitext(filename)[0]
                    staff_member = models.Staff.objects.filter(pin=pin).first()
                    if staff_member:
                        with zip_file.open(filename) as file:
                            new_avatar = ContentFile(file.read())
                            new_avatar.name = filename
                            if new_avatar:
                                if staff_member.avatar:
                                    staff_member.avatar.delete(save=False)
                                staff_member.avatar.save(
                                    new_avatar.name, new_avatar, save=False
                                )
                                staff_member.save()
            logger.info("Staff photos updated successfully")
            messages.success(request, "Фотографии успешно обновлены.")
        except Exception as e:
            logger.error(f"Error processing ZIP file: {str(e)}")
            messages.error(
                request, f"Ошибка при обработке архива с фотографиями: {str(e)}"
            )


class APIKeyCheckView(APIView):
    """
    Проверка API ключа.

    Проверяет наличие и валидность переданного API ключа в заголовке запроса.
    Если ключ отсутствует или недействителен, возвращает соответствующее сообщение об ошибке.

    Методы:
        get: Проверяет API ключ и возвращает данные о его создании и статусе активности.

    Права доступа:
        AllowAny: Доступ открыт для всех пользователей, аутентификация не требуется.
    """

    permission_classes = [AllowAny]

    @swagger_auto_schema(
        operation_summary="Проверка API ключа",
        operation_description="Проверяет наличие и валидность переданного API ключа в заголовке запроса.",
        manual_parameters=[
            openapi.Parameter(
                name="X-API-KEY",
                in_=openapi.IN_HEADER,
                type=openapi.TYPE_STRING,
                required=True,
                description="API ключ для проверки.",
            )
        ],
        responses={
            200: openapi.Response(
                description="Данные о создании и статусе активности API ключа.",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        "data": openapi.Schema(
                            type=openapi.TYPE_OBJECT,
                            properties={
                                "created_at": openapi.Schema(
                                    type=openapi.TYPE_STRING,
                                    format="date-time",
                                    description="Дата и время создания ключа.",
                                ),
                                "is_active": openapi.Schema(
                                    type=openapi.TYPE_BOOLEAN,
                                    description="Статус активности ключа.",
                                ),
                            },
                        )
                    },
                ),
            ),
            400: openapi.Response(
                description="Некорректный запрос: отсутствует или недействителен API ключ.",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        "message": openapi.Schema(
                            type=openapi.TYPE_STRING,
                            description="Сообщение об ошибке.",
                        ),
                        "error": openapi.Schema(
                            type=openapi.TYPE_STRING,
                            description="Описание ошибки.",
                        ),
                    },
                ),
            ),
        },
    )
    def get(self, request, *args, **kwargs):
        """
        Проверяет API ключ и возвращает данные о его создании и статусе активности.

        Аргументы:
            request (HttpRequest): Объект HTTP запроса.

        Возвращает:
            Response: Ответ с данными о создании и статусе активности API ключа, либо сообщение об ошибке.
        """
        logger.info("API Key check request received")

        api_key = request.headers.get("X-API-KEY")

        if not api_key:
            logger.warning("API Key is missing in the request")
            return Response(
                {"message": "API Key is missing"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            secret_key = utils.APIKeyUtility.get_secret_key()
            logger.debug("Secret key retrieved successfully")
            data = utils.APIKeyUtility.decrypt_data(
                api_key, secret_key, fields=("created_at", "is_active")
            )
            logger.info("API Key is valid")
            return Response({"data": data}, status=status.HTTP_200_OK)
        except Exception as e:
            logger.error(f"Invalid API Key: {str(e)}")
            return Response(
                {"message": "Invalid API Key", "error": str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )


def password_reset_request_view(request):
    if request.method == "POST":
        identifier = request.POST.get("identifier")
        ip_address = utils.get_client_ip(request)
        logger.error(str(ip_address))
        user = (
            User.objects.filter(username=identifier).first()
            or User.objects.filter(email=identifier).first()
        )

        user_timezone = utils.get_user_timezone(request)

        if user:
            last_request_time = models.PasswordResetRequestLog.get_last_request_time(
                user, ip_address
            )
            if not models.PasswordResetRequestLog.can_request_again(user, ip_address):
                next_possible_time = timezone.localtime(
                    last_request_time + timezone.timedelta(minutes=5), user_timezone
                )
                last_request_time_local = timezone.localtime(
                    last_request_time, user_timezone
                )
                messages.warning(
                    request,
                    f"Запрос уже был отправлен. Повторный запрос возможен в {next_possible_time.strftime('%H:%M:%S %Z')} ({next_possible_time.tzinfo}). Последний запрос был в {last_request_time_local.strftime('%H:%M:%S %Z')} ({last_request_time_local.tzinfo}).",
                )
            else:
                utils.send_password_reset_email(user, request)
                models.PasswordResetRequestLog.log_request(user, ip_address)
                current_time_local = timezone.localtime(timezone.now(), user_timezone)
                messages.success(
                    request,
                    f"Если пользователь существует, ссылка для сброса пароля была отправлена на его электронную почту. Последний запрос был в {current_time_local.strftime('%H:%M:%S %Z')} ({current_time_local.tzinfo}).",
                )
        else:
            messages.info(
                request,
                "Если пользователь существует, ссылка для сброса пароля была отправлена на его электронную почту.",
            )

        return redirect("password_reset_request")

    return render(request, "password_reset_request.html")


def password_reset_confirm_view(request, token):
    reset_token = get_object_or_404(models.PasswordResetToken, token=token)

    if not reset_token.is_valid():
        messages.error(request, "Этот токен для сброса пароля больше не действителен.")
        return redirect("password_reset_request")

    if request.method == "POST":
        new_password = request.POST.get("password")
        user = reset_token.user
        user.set_password(new_password)
        user.save()

        if models.PasswordResetToken.objects.mark_as_used(token):
            messages.success(
                request,
                "Пароль успешно сброшен. Вы можете войти в систему с новым паролем.",
            )
            return redirect("reac_app")
        else:
            messages.error(request, "Ошибка при обновлении токена. Попробуйте снова.")
            return redirect("password_reset_confirm", token=token)

    return render(request, "password_reset_confirm.html", {"token": token})
