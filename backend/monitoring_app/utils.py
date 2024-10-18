import os
import re
import json
import datetime
from functools import wraps
from typing import Any, Dict, List, Optional
from concurrent.futures import ThreadPoolExecutor

import pytz
import requests
import pandas as pd
from openpyxl import Workbook
from django.urls import reverse
from django.conf import settings
from django.db import transaction
from django.utils import timezone
from rest_framework import status
from django.core.cache import cache
from django.http import HttpRequest
from cryptography.fernet import Fernet
from django.core.mail import send_mail
from django.utils.html import format_html
from rest_framework.response import Response
from django.contrib.admin import SimpleListFilter
from django.utils.translation import gettext_lazy as _
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill

import cv2
import torch
import logging
import numpy as np
import torch.nn as nn
import torch.optim as optim
from torch.utils.data import DataLoader
from insightface.app import FaceAnalysis
from torchvision import models as torch_models
from sklearn.model_selection import train_test_split
from sklearn.metrics.pairwise import cosine_similarity

from monitoring_app import models
from rest_framework.exceptions import ValidationError

DAYS = settings.DAYS

logger = logging.getLogger(__name__)

arcface_model = None


class FaceRecognitionResNet(nn.Module):
    def __init__(self, input_size):
        super(FaceRecognitionResNet, self).__init__()
        self.resnet = torch_models.resnet50(pretrained=True)
        self.resnet.fc = nn.Linear(self.resnet.fc.in_features, 512)
        self.fc1 = nn.Linear(512, 128)
        self.fc2 = nn.Linear(128, 1)

    def forward(self, x):
        x = self.resnet(x)
        x = torch.relu(self.fc1(x))
        x = torch.sigmoid(self.fc2(x))
        return x


def train_face_recognition_model(staff, embeddings, negative_embeddings):
    try:
        device = get_device()

        embeddings_combined = np.vstack([embeddings, negative_embeddings])
        labels = [1] * len(embeddings) + [0] * len(negative_embeddings)

        inputs_train, inputs_val, labels_train, labels_val = train_test_split(
            embeddings_combined, labels, test_size=0.2, random_state=42
        )

        train_data = list(zip(inputs_train, labels_train))
        val_data = list(zip(inputs_val, labels_val))

        train_loader = DataLoader(train_data, batch_size=32, shuffle=True, num_workers=4)
        val_loader = DataLoader(val_data, batch_size=32, shuffle=False, num_workers=4)

        model = FaceRecognitionResNet(input_size=len(embeddings[0]))
        model = nn.DataParallel(model)
        model.to(device)

        criterion = nn.BCELoss()
        optimizer = optim.Adam(model.parameters(), lr=0.001)

        epochs = 20
        for epoch in range(epochs):
            model.train()
            train_loss = 0.0

            for batch_inputs, batch_labels in train_loader:
                batch_inputs = torch.tensor(batch_inputs, dtype=torch.float32).to(device)
                batch_labels = torch.tensor(batch_labels, dtype=torch.float32).to(device)

                optimizer.zero_grad()
                outputs = model(batch_inputs)
                loss = criterion(outputs.squeeze(), batch_labels)
                loss.backward()
                optimizer.step()

                train_loss += loss.item()

            model.eval()
            val_loss = 0.0
            with torch.no_grad():
                for batch_inputs, batch_labels in val_loader:
                    batch_inputs = torch.tensor(batch_inputs, dtype=torch.float32).to(device)
                    batch_labels = torch.tensor(batch_labels, dtype=torch.float32).to(device)

                    outputs = model(batch_inputs)
                    loss = criterion(outputs.squeeze(), batch_labels)
                    val_loss += loss.item()

            print(
                f'Epoch {epoch+1}/{epochs}, Train Loss: {train_loss/len(train_loader)}, Validation Loss: {val_loss/len(val_loader)}'
            )

        model_path = os.path.join(
            os.path.dirname(staff.avatar.path), f'{staff.pin}_resnet_model.pkl'
        )
        torch.save(model.state_dict(), model_path)
        logger.info(f"ResNet model for {staff.pin} saved at {model_path}")

    except Exception as e:
        logger.error(f"Error training model for {staff.pin}: {str(e)}")
        raise e


def load_model_for_staff(staff):
    """
    Загружает ранее обученную нейронную сеть для указанного сотрудника.

    Args:
        staff (Staff): Объект сотрудника, для которого загружается модель.

    Returns:
        FaceRecognitionNN: Загрузка модели нейросети для использования.

    Raises:
        ValueError: Если модель не найдена.
    """
    model_path = os.path.join(os.path.dirname(staff.avatar.path), f'{staff.pin}_model.pkl')
    if not os.path.exists(model_path):
        raise ValueError(f"Модель для {staff.pin} не найдена")

    model = FaceRecognitionResNet(input_size=512)
    model.load_state_dict(torch.load(model_path))
    model.eval()
    return model


def load_arcface_model():
    """
    Загружает модель ArcFace для распознавания лиц.

    Инициализирует модель ArcFace, если она еще не загружена.
    """
    global arcface_model
    if arcface_model is None:
        arcface_model = FaceAnalysis(providers=['CUDAExecutionProvider', 'CPUExecutionProvider'])
        arcface_model.prepare(ctx_id=0, det_size=(640, 640))
        logger.info("ArcFace model loaded and ready.")


def get_device():
    """
    Определяет доступное устройство (GPU или CPU).

    Использует CUDA, если оно доступно, в противном случае использует CPU.

    Returns:
        torch.device: Возвращает устройство для выполнения операций (GPU или CPU).
    """
    return torch.device("cuda" if torch.cuda.is_available() else "cpu")


def load_image_from_memory(file):
    """
    Загружает изображение из InMemoryUploadedFile и конвертирует его в numpy array.

    Использует OpenCV для чтения изображения из загруженного файла. Преобразует
    изображение в формат numpy array для дальнейшей обработки.
    Args:
        file (InMemoryUploadedFile): Файл, загруженный через Django.

    Returns:
        numpy.ndarray: Изображение в виде массива numpy.

    Raises:
        ValidationError: Если не удается прочитать изображение.
    """
    try:
        file_bytes = np.frombuffer(file.read(), np.uint8)
        image = cv2.imdecode(file_bytes, cv2.IMREAD_COLOR)
        if image is None:
            raise ValidationError("Невозможно прочитать изображение.")
        return image
    except Exception as e:
        logger.error(f"Ошибка при чтении изображения: {e}")
        raise ValidationError(f"Ошибка чтения изображения: {str(e)}")


def create_face_encoding(image_file):
    """
    Создает face encoding (закодированное лицо) для изображения с использованием ArcFace.

    Args:
        image_file (str, numpy.ndarray или InMemoryUploadedFile): Путь к изображению, numpy array или загруженный файл.

    Returns:
        list: Закодированное лицо (embedding) в виде списка чисел.

    Raises:
        ValidationError: Если не удается создать encoding для изображения.
    """
    try:
        load_arcface_model()

        if isinstance(image_file, str):
            img = cv2.imread(image_file)

        elif isinstance(image_file, np.ndarray):
            img = image_file

        else:
            img = load_image_from_memory(image_file)

        faces = arcface_model.get(img)
        if not faces:
            raise ValidationError("Лицо не найдено на изображении")

        return faces[0].embedding.tolist()

    except Exception as e:
        logger.error(f"Ошибка при создании encoding: {e}")
        raise ValidationError(f"Ошибка создания encoding: {str(e)}")


def compare_face(staff, image_file):
    """
    Сравнивает лицо сотрудника с новым изображением, используя косинусное расстояние.

    Эта функция получает ранее сохраненное закодированное лицо сотрудника и сравнивает
    его с закодированным лицом, созданным на основе нового изображения. Для сравнения
    используется косинусное расстояние.

    Args:
        staff (Staff): Сотрудник, чье лицо сравнивается.
        image_file (InMemoryUploadedFile): Новое изображение для сравнения.

    Returns:
        tuple: (verified, cosine_distance):
            - verified (bool): Результат сравнения (True, если лица совпадают).
            - cosine_distance (float): Фактическое косинусное расстояние между эмбеддингами.

    Raises:
        ValidationError: Если возникает ошибка при сравнении лиц.
    """
    try:
        stored_mask = np.array(staff.face_mask.mask_encoding)
        new_encoding = create_face_encoding(image_file)

        cosine_distance = cosine_similarity([stored_mask], [new_encoding])[0][0]

        cosine_distance = (1 + cosine_distance) / 2

        threshold = settings.FACE_RECOGNITION_THRESHOLD

        verified = cosine_distance > threshold

        return verified, cosine_distance
    except Exception as e:
        logger.error(f"Ошибка при сравнении лица: {e}")
        raise ValidationError(f"Ошибка сравнения лица: {str(e)}")


def compare_face_with_nn(staff, image_file):
    """
    Сравнивает лицо сотрудника с новым изображением, используя обученную нейронную сеть.

    Использует предобученную нейронную сеть для проверки совпадения лиц.

    Args:
        staff (Staff): Сотрудник, чье лицо сравнивается.
        image_file (InMemoryUploadedFile): Новое изображение для сравнения.

    Returns:
        tuple: (verified, distance):
            - verified (bool): True, если лица совпадают.
            - distance (float): Косинусное расстояние между эмбеддингами.

    Raises:
        ValidationError: Если возникает ошибка при сравнении лиц.
    """
    try:
        load_arcface_model()

        new_encoding = create_face_encoding(image_file)

        model_path = os.path.join(os.path.dirname(staff.avatar.path), f'{staff.pin}_model.pkl')
        device = get_device()

        if not os.path.exists(model_path):
            raise ValueError(f"Модель для {staff.pin} не найдена")

        model = FaceRecognitionResNet()(input_size=len(new_encoding)).to(device)
        model.load_state_dict(torch.load(model_path))
        model.eval()

        input_tensor = torch.tensor([new_encoding], dtype=torch.float32).to(device)
        output = model(input_tensor).item()

        threshold = settings.FACE_RECOGNITION_THRESHOLD
        verified = output > threshold

        logger.info(f"Face comparison using NN for {staff.pin}: cosine distance = {output}")

        return verified, output

    except Exception as e:
        logger.error(f"Ошибка при сравнении лица с нейронной сетью для {staff.pin}: {str(e)}")
        raise ValidationError(
            f"Ошибка при сравнении лица с нейронной сетью для {staff.pin}: {str(e)}"
        )


def get_client_ip(request):
    """Get the client IP address from the request, considering proxy setups."""
    x_forwarded_for = request.META.get("HTTP_X_FORWARDED_FOR")
    if x_forwarded_for:
        ip = x_forwarded_for.split(",")[0].strip()
    else:
        ip = request.META.get("REMOTE_ADDR")
    return ip


def format_duration(duration_seconds):
    """Converts a duration in seconds to a more human-readable format."""
    if duration_seconds < 60:
        return f"{duration_seconds:.2f} seconds"
    elif duration_seconds < 3600:
        minutes = duration_seconds // 60
        seconds = duration_seconds % 60
        return f"{minutes:.0f} minute(s) {seconds:.2f} seconds"
    else:
        hours = duration_seconds // 3600
        minutes = (duration_seconds % 3600) // 60
        seconds = duration_seconds % 60
        return f"{hours:.0f} hour(s) {minutes:.0f} minute(s) {seconds:.2f} seconds"


class HierarchicalDepartmentFilter(SimpleListFilter):
    title = _("Department")
    parameter_name = "staff_department"

    def lookups(self, request, model_admin):
        departments = models.ChildDepartment.objects.all()
        return [(dept.id, dept.name) for dept in departments]

    def queryset(self, request, queryset):
        if self.value():
            department = models.ChildDepartment.objects.get(pk=self.value())
            child_departments = department.get_all_child_departments()
            child_department_ids = [dept.id for dept in child_departments] + [department.id]
            return queryset.filter(staff__department__in=child_department_ids)
        return queryset


class APIKeyUtility:
    _secret_key = None

    @staticmethod
    def get_secret_key():
        if APIKeyUtility._secret_key is None:
            if not settings.SECRET_API:
                secret_key = Fernet.generate_key().decode("utf-8")

                dotenv = settings.DOTENV_PATH

                with open(dotenv, mode="ab") as f:
                    f.write(f"""\nSECRET_API={secret_key}\n""".encode("utf-8"))

                APIKeyUtility._secret_key = secret_key
            else:
                APIKeyUtility._secret_key = settings.SECRET_API

        return APIKeyUtility._secret_key

    @staticmethod
    def encrypt_data(data, secret_key):
        f = Fernet(secret_key.encode())
        encrypted_data = f.encrypt(json.dumps(data).encode())
        return encrypted_data.decode()

    @staticmethod
    def decrypt_data(encrypted_data, secret_key, fields=("is_active",)):
        f = Fernet(secret_key.encode())
        decrypted_data = f.decrypt(encrypted_data.encode())
        data = json.loads(decrypted_data.decode())

        return {field: data.get(field) for field in fields}

    @staticmethod
    def generate_api_key(key_name, created_by):
        secret_key = APIKeyUtility.get_secret_key()
        data = {
            "key_name": key_name,
            "created_by": created_by.username,
            "created_at": timezone.now().isoformat(),
            "is_activate": True,
        }
        encrypted_data = APIKeyUtility.encrypt_data(data, secret_key)
        return encrypted_data, secret_key


def get_attendance_data(pin: str):
    """
    Получает данные о посещаемости сотрудника за предыдущий день.

    Args:
        pin (str): Идентификационный номер сотрудника (PIN-код).

    Returns:
        list: Список словарей, содержащих информацию о каждом проходе
              сотрудника через систему контроля доступа за указанный день.
              Если произошла ошибка, возвращает пустой список.
    """
    prev_date = datetime.datetime.now() - datetime.timedelta(days=DAYS)

    eventTime_first = prev_date.strftime("%Y-%m-%d 00:00:00")
    eventTime_last = prev_date.strftime("%Y-%m-%d 23:59:59")
    access_token = settings.API_KEY

    try:
        params = {
            "endDate": eventTime_last,
            "pageNo": "1",
            "pageSize": "1000",
            "personPin": pin,
            "startDate": eventTime_first,
            "access_token": access_token,
        }
        response = requests.get(
            settings.API_URL + "/api/transaction/listAttTransaction",
            params=params,
            timeout=10,
        )
        response.raise_for_status()
        response_json = response.json()
        return response_json.get("data", [])
    except Exception as e:
        print(f"Error, Pin: {pin}, Error: {e}")
        return []


def get_all_attendance():
    """
    Получает данные о посещаемости всех сотрудников за текущий день
    и сохраняет их в базе данных.

    Выполняет запрос данных о посещаемости для каждого сотрудника
    параллельно с помощью ThreadPoolExecutor. Затем для каждого
    сотрудника извлекает первое и последнее время прохода
    (если данные о посещаемости имеются) и сохраняет их
    в модели StaffAttendance.

    Returns:
        None
    """
    pins = models.Staff.objects.values_list("pin", flat=True)
    attendance_data = {}

    with ThreadPoolExecutor(max_workers=6) as executor:
        for pin, data in zip(pins, executor.map(get_attendance_data, pins)):
            attendance_data[pin] = data

    prev_date = timezone.now() - timezone.timedelta(days=DAYS)
    next_day = prev_date + timezone.timedelta(days=1)

    updates = []

    for pin, data in attendance_data.items():
        staff = models.Staff.objects.get(pin=pin)
        if data:
            first_event = data[-1]
            last_event = data[0] if len(data) > 1 else first_event

            first_event_time = timezone.make_aware(
                timezone.datetime.fromisoformat(first_event["eventTime"])
            )
            last_event_time = (
                timezone.make_aware(timezone.datetime.fromisoformat(last_event["eventTime"]))
                if len(data) > 1
                else first_event_time
            )

            first_area_name = first_event.get("areaName")
            last_area_name = last_event.get("areaName")

            area_name_in = first_area_name or "Unknown"
            area_name_out = last_area_name or "Unknown"

        else:
            first_event_time = None
            last_event_time = None
            area_name_in = "Unknown"
            area_name_out = "Unknown"

        date_at = next_day.date()

        attendance, created = models.StaffAttendance.objects.get_or_create(
            staff=staff,
            date_at=date_at,
            defaults={
                "first_in": first_event_time,
                "last_out": last_event_time,
                "area_name_in": area_name_in,
                "area_name_out": area_name_out,
            },
        )
        if not created:
            attendance.first_in = first_event_time
            attendance.last_out = last_event_time
            attendance.area_name_in = area_name_in
            attendance.area_name_out = area_name_out
            updates.append(attendance)

    with transaction.atomic():
        if updates:
            models.StaffAttendance.objects.bulk_update(
                updates, ["first_in", "last_out", "area_name_in", "area_name_out"]
            )


def password_check(password: str) -> bool:
    """
    Проверяет, соответствует ли пароль требованиям сложности системы.

    Эта функция проверяет строку пароля на основе следующих критериев:

        - Минимальная длина 8 символов.
        - Содержит хотя бы одну заглавную букву (A-Z)
        - Содержит хотя бы одну строчную букву (a-z)
        - Содержит хотя бы одну цифру (0-9)
        - Содержит хотя бы один специальный символ из следующего набора: #?!@$%^&*-

    Args:
        пароль (str): строка пароля, которую необходимо проверить.

    Returns:
        bool: True, если пароль соответствует всем требованиям сложности, в противном случае — False
    """
    return bool(
        re.match(r"^(?=.*?[A-Z])(?=.*?[a-z])(?=.*?[0-9])(?=.*?[#?!@$%^&*-]).{8,}$", password)
    )


def fetch_data(url: str) -> Dict[str, Any]:
    try:
        response = requests.get(url, timeout=20)
        response.raise_for_status()
        return response.json()
    except requests.RequestException as e:
        print(f"Error fetching data: {e}")
        return {}


def parse_attendance_data(data: List[Dict[str, Any]]) -> List[List[Optional[str]]]:
    rows: List[List[Optional[str]]] = []
    timezone_pattern = re.compile(r"\+\d{2}:\d{2}")
    holidays_cache = cache.get("holidays_cache")

    if not holidays_cache:
        holidays_cache = {holiday.date: holiday for holiday in models.PublicHoliday.objects.all()}
        cache.set("holidays_cache", holidays_cache, timeout=1 * 12 * 60)

    def parse_datetime_with_timezone(dt_str: Optional[str]) -> Optional[str]:
        if not dt_str:
            return None
        match = timezone_pattern.search(dt_str)
        if match:
            timezone = match.group(0)
            dt_format = f"%Y-%m-%dT%H:%M:%S{timezone}"
            return datetime.datetime.strptime(dt_str, dt_format).strftime("%H:%M:%S")
        return None

    for record in data:
        for date, details in record.items():
            try:
                date_obj = datetime.datetime.strptime(date, "%Y-%m-%d")
                date_str = date_obj.strftime("%d.%m.%Y")
            except ValueError as e:
                print(f"Error parsing date: {e}")
                continue

            public_holiday = holidays_cache.get(date_obj)
            is_holiday = public_holiday and not public_holiday.is_working_day

            department = details.get("department", "").replace("_", " ").capitalize()
            for attendance in details.get("attendance", []):
                try:
                    staff_fio = attendance.get("staff_fio", "")
                    first_in = parse_datetime_with_timezone(attendance.get("first_in"))
                    last_out = parse_datetime_with_timezone(attendance.get("last_out"))

                    if date_obj.weekday() >= 5 or is_holiday:
                        if first_in and last_out:
                            attendance_info = f"{first_in} - {last_out}"
                        else:
                            attendance_info = "Выходной"
                    else:
                        attendance_info = (
                            f"{first_in} - {last_out}" if first_in and last_out else "Отсутствие"
                        )

                    rows.append([staff_fio, department, date_str, attendance_info])
                except KeyError as e:
                    print(f"Missing expected key: {e}")
                except Exception as e:
                    print(f"Error processing record: {e}")
    return rows


def create_dataframe(rows: List[List[Optional[str]]]) -> pd.DataFrame:
    df = pd.DataFrame(rows, columns=["ФИО", "Отдел", "Дата", "Посещаемость"])
    df_pivot = df.pivot_table(
        index=["ФИО", "Отдел"],
        columns="Дата",
        values="Посещаемость",
        aggfunc="first",
        fill_value="Отсутствие",
    )
    df_pivot_sorted = df_pivot.reindex(
        sorted(
            df_pivot.columns,
            key=lambda x: pd.to_datetime(x, format="%d.%m.%Y"),
            reverse=True,
        ),
        axis=1,
    )
    return df_pivot_sorted


def save_to_excel(df_pivot_sorted: pd.DataFrame) -> Workbook:
    wb = Workbook()
    ws = wb.active

    data_font = Font(name="Roboto", size=11)
    data_alignment = Alignment(horizontal="center", vertical="center")

    absence_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    absence_font = Font(color="FFFFFF")

    df_flat = df_pivot_sorted.reset_index()
    df_flat_sorted = df_flat.sort_values(by=["Отдел", "ФИО"])

    max_col_widths = [0] * len(df_flat_sorted.columns)
    max_row_heights = [0] * (len(df_flat_sorted) + 1)

    for r_idx, r in enumerate(dataframe_to_rows(df_flat_sorted, index=False, header=True), 1):
        for c_idx, value in enumerate(r, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            if value == "Отсутствие":
                cell.fill = absence_fill
                cell.font = absence_font

            value_length = len(str(value))

            if value_length > max_col_widths[c_idx - 1]:
                max_col_widths[c_idx - 1] = value_length

            if value_length > max_row_heights[r_idx - 1]:
                max_row_heights[r_idx - 1] = value_length

    header_font = Font(name="Roboto", size=14, bold=True)
    for cell in ws[1]:
        cell.font = header_font

    for idx, col_width in enumerate(max_col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = col_width + 2

    for idx, row_height in enumerate(max_row_heights, 1):
        ws.row_dimensions[idx].height = row_height * 3

    return wb


def add_api_key_header(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        request = args[0]

        if isinstance(request, HttpRequest):
            api_key = models.APIKey.objects.filter(is_active=True).first()
            if api_key:
                request.META["HTTP_X_API_KEY"] = api_key.key
            else:
                return Response(
                    {"error": "No active API key available for authentication."},
                    status=status.HTTP_403_FORBIDDEN,
                )

        return func(*args, **kwargs)

    return wrapper


def send_password_reset_email(user, request):
    reset_link = f"{request.scheme}://{request.get_host()}{reverse('password_reset_confirm', args=[models.PasswordResetToken.generate_token(user)])}"

    subject = "Инструкция по сбросу пароля"
    html_message = format_html(
        """
        <div style="font-family: 'Segoe UI', 'Roboto', 'Helvetica Neue', 'Arial', sans-serif; max-width: 600px; margin: auto; padding: 20px; border: 1px solid #ececec; border-radius: 10px; background-color: #ffffff;">
            <h2 style="color: #007bff; text-align: center; margin-bottom: 20px;">Сброс пароля</h2>
            <p style="color: #333; line-height: 1.6; margin-bottom: 20px; text-align: center;">
                Здравствуйте, {username}. Вы запросили сброс пароля для своего аккаунта. 
                Пожалуйста, нажмите кнопку ниже, чтобы сбросить пароль.
            </p>
            <div style="text-align: center; margin-bottom: 30px;">
                <a href="{reset_link}" style="display: inline-block; padding: 15px 30px; color: #ffffff; background-color: #007bff; text-decoration: none; border-radius: 5px; font-size: 18px; font-weight: bold;">
                    Сбросить пароль
                </a>
            </div>
            <p style="color: #555; line-height: 1.6; margin-bottom: 20px; text-align: center;">
                Ссылка для сброса пароля действительна в течение <strong>1 часа</strong>. Если вы не запрашивали сброс пароля, просто проигнорируйте это письмо.
            </p>
            <hr style="border: none; border-top: 1px solid #ececec; margin: 30px 0;">
            <p style="font-size: 12px; color: #999; text-align: center;">
                Мы рады помочь вам сохранить безопасность вашего аккаунта. 
            </p>
        </div>
        """,
        username=user.username,
        reset_link=reset_link,
    )

    plain_message = (
        f"Здравствуйте, {user.username}!\n\n"
        "Вы получили это письмо, потому что был запрошен сброс пароля для вашего аккаунта.\n\n"
        f"Для сброса пароля перейдите по следующей ссылке: {reset_link}\n\n"
        "Эта ссылка будет действительна в течение 1 часа. Для вашей безопасности не передавайте эту ссылку другим лицам.\n\n"
        "Если вы не запрашивали сброс пароля, проигнорируйте это письмо, и ваш пароль останется неизменным.\n\n"
        "С уважением,\n"
        "Команда поддержки."
    )

    send_mail(
        subject,
        plain_message,
        settings.DEFAULT_FROM_EMAIL,
        [user.email],
        html_message=html_message,
    )


def get_user_timezone(request):
    user_timezone = request.session.get("timezone")
    if not user_timezone:
        user_timezone = settings.TIME_ZONE
    return pytz.timezone(user_timezone)


def normalize_id(department_id):
    """
    Нормализует ID отдела, удаляя ведущие нули, если ID состоит только из цифр.
    Если ID содержит буквы, он остаётся без изменений.

    Args:
        department_id (str): ID отдела.

    Returns:
        str: Нормализованный ID.
    """
    if department_id.isdigit():
        return str(int(department_id))
    return department_id


def transliterate(name):
    slovar = {
        "а": "a",
        "б": "b",
        "в": "v",
        "г": "g",
        "д": "d",
        "е": "e",
        "ё": "yo",
        "ж": "zh",
        "з": "z",
        "и": "i",
        "й": "y",
        "к": "k",
        "л": "l",
        "м": "m",
        "н": "n",
        "о": "o",
        "п": "p",
        "р": "r",
        "с": "s",
        "т": "t",
        "у": "u",
        "ф": "f",
        "х": "h",
        "ц": "ts",
        "ч": "ch",
        "ш": "sh",
        "щ": "sch",
        "ъ": "",
        "ы": "y",
        "ь": "",
        "э": "e",
        "ю": "yu",
        "я": "ya",
        " ": " ",
        "-": "-",
        ".": ".",
        ",": ",",
        "!": "!",
        "?": "?",
        ":": ":",
    }

    name = name.lower()
    translit = []
    for letter in name:
        translit.append(slovar.get(letter, letter))

    return ''.join(translit)
