import re
import requests
import pandas as pd
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import Font, Alignment
from typing import List, Dict, Any, Optional
from concurrent.futures import ThreadPoolExecutor
from openpyxl.utils.dataframe import dataframe_to_rows


def fetch_data(url: str) -> Dict[str, Any]:
    try:
        response = requests.get(url)
        response.raise_for_status()
        return response.json()
    except requests.RequestException as e:
        print(f"Error fetching data: {e}")
        return {}


def parse_attendance_data(data: Dict[str, Any]) -> List[List[Optional[str]]]:
    rows: List[List[Optional[str]]] = []
    try:
        for attendance in data.get("attendance", []):
            for date, records in attendance.items():
                date_obj = datetime.strptime(date, "%Y-%m-%d")
                date_str = date_obj.strftime("%d.%m.%Y")

                for record in records:
                    department_name = data.get("department_name", "")
                    staff_fio = record.get("staff_fio", "")

                    timezone_pattern = re.compile(r"\+\d{2}:\d{2}")

                    def parse_datetime_with_timezone(
                        dt_str: Optional[str],
                    ) -> Optional[str]:
                        if not dt_str:
                            return None
                        match = timezone_pattern.search(dt_str)
                        if match:
                            timezone = match.group(0)
                            dt_format = f"%Y-%m-%dT%H:%M:%S{timezone}"
                            return datetime.strptime(dt_str, dt_format).strftime(
                                "%H:%M:%S"
                            )
                        return None

                    first_in = parse_datetime_with_timezone(record.get("first_in"))
                    last_out = parse_datetime_with_timezone(record.get("last_out"))

                    if date_obj.weekday() < 5:
                        if not first_in and not last_out:
                            attendance_info = "Отсутствие"
                        else:
                            attendance_info = (
                                f"{first_in} - {last_out}"
                                if first_in and last_out
                                else "Отсутствие"
                            )
                    else:
                        if first_in and last_out:
                            attendance_info = f"{first_in} - {last_out}"
                        else:
                            attendance_info = "Выходной"

                    rows.append([staff_fio, date_str, attendance_info])
    except KeyError as e:
        print(f"Missing expected key: {e}")
        return None
    except Exception as e:
        print(f"Error parsing data: {e}")
        return None
    return rows


def create_dataframe(rows: List[List[Optional[str]]]) -> pd.DataFrame:
    df = pd.DataFrame(rows, columns=["ФИО", "Дата", "Посещаемость"])
    df_pivot = df.pivot_table(
        index=["ФИО"],
        columns="Дата",
        values="Посещаемость",
        aggfunc="first",
        fill_value="Отсутствие",
    )
    df_pivot_sorted = df_pivot.reindex(
        sorted(
            df_pivot.columns,
            key=lambda x: pd.to_datetime(x, format="%d.%m.%Y"),
            reverse=True,
        ),
        axis=1,
    )
    return df_pivot_sorted


def save_to_excel(df_pivot_sorted: pd.DataFrame, filename: str) -> None:
    wb = Workbook()
    ws = wb.active

    data_font = Font(name="Roboto", size=10)
    data_alignment = Alignment(horizontal="center", vertical="center")

    for r_idx, r in enumerate(
        dataframe_to_rows(df_pivot_sorted, index=True, header=True), 1
    ):
        for c_idx, value in enumerate(r, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment

    header_font = Font(name="Roboto", size=12, bold=True)
    for cell in ws[1]:
        cell.font = header_font

    wb.save(filename)


def main() -> None:
    url = "http://localhost:8000/api/department/stats/4958/?end_date=2024-05-29&start_date=2024-02-01"
    data = fetch_data(url)

    with ThreadPoolExecutor() as executor:
        future = executor.submit(parse_attendance_data, data)
        rows = future.result()

    if rows:
        df_pivot_sorted = create_dataframe(rows)
        department_name = data.get("department_name", "Отдел").replace(" ", "_")
        filename = f"Посещаемость_{department_name}.xlsx"
        save_to_excel(df_pivot_sorted, filename)


if __name__ == "__main__":
    main()
